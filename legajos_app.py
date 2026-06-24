import os
import subprocess
import tempfile
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import date

from openpyxl import load_workbook
from PIL import Image, ImageTk
from pypdf import PdfWriter
from tkcalendar import DateEntry

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    BaseApp = TkinterDnD.Tk
    DND_OK = True
except ImportError:
    BaseApp = tk.Tk
    DND_OK = False


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCRIPT_DEFAULT = os.path.join(BASE_DIR, "legajos_core.py")
MAESTRA_DEFAULT = os.path.join(BASE_DIR, "MaestraEstaciones.xlsx")

MATRICULA_DOC_TYPES = [
    "ACTA_INSTALACION",
    "FICHA_MATRICULA",
    "INFORME_MATRICULA",
    "INFORME_INSTALACION",
    "CONVENIO",
    "ANEXO",
    "OTRO",
]

CATEGORIAS_ESTACION = [
    "MANTENIMIENTO",
    "CHECKLIST_MANTENIMIENTO",
    "ESTADO_SITUACIONAL",
    "INSPECCION",
    "CALIBRACION",
    "CLASIFICACION_EMPLAZAMIENTO",
    "CERTIFICADO_COMPROBACION_SENSOR",
    "AFOROS",
    "CALIDAD_DATOS",
    "INCIDENCIAS",
    "INSTALACIONES_NUEVAS",
    "INSTALACION_EQUIPO",
    "CESE_OBSERVADOR",
    "SINIESTROS",
    "REUBICACION",
    "SOLICITUD_REUBICACION",
    "BAJA_ESTACION",
    "CLAUSURA_ESTACION",
    "ACTUALIZACION_METADATA",
]

SINIESTRO_SUBTIPOS = [
    "ROBO",
    "DAÑO_POR_TERCEROS",
    "DAÑO_POR_FENOMENO_NATURAL",
    "ACCIDENTE",
    "OTRO",
]

SENSOR_VARIABLES = [
    "TEMPERATURA",
    "HUMEDAD_RELATIVA",
    "PRECIPITACION",
    "RADIACION_SOLAR",
    "RADIACION_UV",
    "VIENTO",
    "PRESION_ATMOSFERICA",
    "OTRO",
]


IOARR_DOC_TYPES = [
    "NOTA_ELEVACION_DHI",
    "INFORME_DHI",
    "NOTA_ELEVACION_SGR",
    "INFORME_SGR",
    "INFORME_DRD",
    "INFORME_TECNICO",
    "FORMATO_07C",
    "PRESUPUESTO",
    "OTRO",
]

CONVENIO_DZ_DOC_TYPES = [
    "CONVENIO",
    "CONVENIO_MARCO",
    "ADENDA",
    "ACTA_TRANSFERENCIA_FISICA",
    "ACTA_REUNION",
    "INFORME_SUSTENTO",
    "OFICIO",
    "MEMORANDO",
    "OTRO",
]


def slug_text(text):
    text = str(text or "").strip()
    repl = {"Á":"A","É":"E","Í":"I","Ó":"O","Ú":"U","á":"a","é":"e","í":"i","ó":"o","ú":"u","Ñ":"N","ñ":"n"}
    for a, b in repl.items():
        text = text.replace(a, b)
    cleaned = []
    for ch in text:
        if ch.isalnum() or ch in "-_":
            cleaned.append(ch)
        elif ch.isspace():
            cleaned.append("_")
    out = "".join(cleaned).strip("_")
    while "__" in out:
        out = out.replace("__", "_")
    return out or "SIN_NOMBRE"


class DropField(ttk.Frame):
    def __init__(self, master, label, is_dir=False, filetypes=None):
        super().__init__(master)
        self.is_dir = is_dir
        self.filetypes = filetypes or [("Todos los archivos", "*.*")]
        self.var = tk.StringVar()

        ttk.Label(self, text=label, width=22).grid(row=0, column=0, padx=5, pady=5, sticky="w")

        self.entry = ttk.Entry(self, textvariable=self.var, width=90)
        self.entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        self.button = ttk.Button(self, text="Seleccionar", command=self.pick)
        self.button.grid(row=0, column=2, padx=5, pady=5)

        self.columnconfigure(1, weight=1)

        if DND_OK:
            self.entry.drop_target_register(DND_FILES)
            self.entry.dnd_bind("<<Drop>>", self.on_drop)

    def on_drop(self, event):
        value = event.data.strip()
        if value.startswith("{") and value.endswith("}"):
            value = value[1:-1]
        value = os.path.realpath(value)
        self.var.set(value)

    def pick(self):
        if self.is_dir:
            path = filedialog.askdirectory()
        else:
            path = filedialog.askopenfilename(filetypes=self.filetypes)

        if path:
            path = os.path.realpath(path)
            self.var.set(path)
            if not os.path.exists(path):
                messagebox.showerror("Ruta inválida", f"La ruta seleccionada no existe:\n{path}")

    def get(self):
        return self.var.get().strip()

    def set(self, value):
        self.var.set(value)


class MultiPhotoRow(ttk.Frame):
    def __init__(self, master, idx, remove_callback, move_up_callback, move_down_callback, on_change_callback):
        super().__init__(master)
        self.idx = idx
        self.remove_callback = remove_callback
        self.move_up_callback = move_up_callback
        self.move_down_callback = move_down_callback
        self.on_change_callback = on_change_callback
        self.var = tk.StringVar()
        self.preview_image = None

        self.var.trace_add("write", self._on_var_change)

        self.label = ttk.Label(self, text=f"Foto {idx}:", width=10)
        self.label.grid(row=0, column=0, padx=5, pady=4, sticky="nw")

        self.entry = ttk.Entry(self, textvariable=self.var, width=60)
        self.entry.grid(row=0, column=1, padx=5, pady=4, sticky="ew")

        self.btn_select = ttk.Button(self, text="Seleccionar", command=self.pick)
        self.btn_select.grid(row=0, column=2, padx=3, pady=4)

        self.btn_up = ttk.Button(self, text="↑", width=3, command=lambda: self.move_up_callback(self))
        self.btn_up.grid(row=0, column=3, padx=2, pady=4)

        self.btn_down = ttk.Button(self, text="↓", width=3, command=lambda: self.move_down_callback(self))
        self.btn_down.grid(row=0, column=4, padx=2, pady=4)

        self.btn_remove = ttk.Button(self, text="Quitar", command=self.remove_me)
        self.btn_remove.grid(row=0, column=5, padx=5, pady=4)

        self.preview_label = ttk.Label(self, text="Sin miniatura")
        self.preview_label.grid(row=1, column=1, padx=5, pady=(0, 6), sticky="w")

        self.columnconfigure(1, weight=1)

        if DND_OK:
            self.entry.drop_target_register(DND_FILES)
            self.entry.dnd_bind("<<Drop>>", self.on_drop)

    def _on_var_change(self, *args):
        self.update_preview()
        self.on_change_callback()

    def on_drop(self, event):
        value = event.data.strip()

        if value.startswith("{") and value.endswith("}"):
            value = value[1:-1]

        if "} {" in value:
            value = value.split("} {")[0].strip("{}")

        value = os.path.realpath(value)
        self.var.set(value)

    def pick(self):
        path = filedialog.askopenfilename(
            filetypes=[
                ("Imágenes y PDF", "*.jpg *.jpeg *.png *.bmp *.webp *.tif *.tiff *.pdf"),
                ("Imágenes", "*.jpg *.jpeg *.png *.bmp *.webp *.tif *.tiff"),
                ("PDF", "*.pdf"),
                ("Todos", "*.*"),
            ]
        )
        if path:
            self.var.set(os.path.realpath(path))

    def remove_me(self):
        self.remove_callback(self)

    def get(self):
        return self.var.get().strip()

    def set_index(self, idx):
        self.idx = idx
        self.label.configure(text=f"Foto {idx}:")

    def update_preview(self):
        path = self.get()
        if not path or not os.path.isfile(path):
            self.preview_image = None
            self.preview_label.configure(image="", text="Sin miniatura")
            return

        ext = os.path.splitext(path)[1].lower()
        if ext == ".pdf":
            self.preview_image = None
            self.preview_label.configure(image="", text="PDF cargado")
            return

        try:
            img = Image.open(path)
            img.thumbnail((140, 140))
            self.preview_image = ImageTk.PhotoImage(img)
            self.preview_label.configure(image=self.preview_image, text="")
        except Exception:
            self.preview_image = None
            self.preview_label.configure(image="", text="No se pudo cargar")


class MatriculaDocRow(ttk.Frame):
    def __init__(self, master, idx, remove_callback, refresh_callback):
        super().__init__(master)
        self.idx = idx
        self.remove_callback = remove_callback
        self.refresh_callback = refresh_callback

        self.path_var = tk.StringVar()
        self.tipo_var = tk.StringVar(value="ACTA_INSTALACION")
        self.nombre_final_var = tk.StringVar()

        self.path_var.trace_add("write", lambda *args: self.refresh_callback())
        self.tipo_var.trace_add("write", lambda *args: self.refresh_callback())

        self.label = ttk.Label(self, text=f"Doc {idx}:", width=8)
        self.label.grid(row=0, column=0, padx=5, pady=4, sticky="w")

        self.entry = ttk.Entry(self, textvariable=self.path_var, width=52)
        self.entry.grid(row=0, column=1, padx=5, pady=4, sticky="ew")

        self.btn_select = ttk.Button(self, text="Seleccionar", command=self.pick)
        self.btn_select.grid(row=0, column=2, padx=3, pady=4)

        ttk.Label(self, text="Tipo:").grid(row=0, column=3, padx=(12, 3), pady=4, sticky="w")
        self.tipo_combo = ttk.Combobox(self, textvariable=self.tipo_var, values=MATRICULA_DOC_TYPES, state="readonly", width=22)
        self.tipo_combo.grid(row=0, column=4, padx=3, pady=4, sticky="w")

        self.btn_remove = ttk.Button(self, text="Quitar", command=lambda: self.remove_callback(self))
        self.btn_remove.grid(row=0, column=5, padx=5, pady=4)

        ttk.Label(self, text="Nombre final:").grid(row=1, column=0, padx=5, pady=(0, 6), sticky="w")
        self.nombre_entry = ttk.Entry(self, textvariable=self.nombre_final_var, width=95)
        self.nombre_entry.grid(row=1, column=1, columnspan=5, padx=5, pady=(0, 6), sticky="ew")

        self.columnconfigure(1, weight=1)

        if DND_OK:
            self.entry.drop_target_register(DND_FILES)
            self.entry.dnd_bind("<<Drop>>", self.on_drop)

    def on_drop(self, event):
        value = event.data.strip()
        if value.startswith("{") and value.endswith("}"):
            value = value[1:-1]
        if "} {" in value:
            value = value.split("} {")[0].strip("{}")
        self.path_var.set(os.path.realpath(value))

    def pick(self):
        path = filedialog.askopenfilename(filetypes=[("PDF y documentos", "*.pdf *.xlsx *.xls *.doc *.docx *.jpg *.jpeg *.png"), ("Todos", "*.*")])
        if path:
            self.path_var.set(os.path.realpath(path))

    def get_path(self):
        return self.path_var.get().strip()

    def get_tipo(self):
        return self.tipo_var.get().strip() or "OTRO"

    def get_nombre_final(self):
        return self.nombre_final_var.get().strip()

    def set_index(self, idx):
        self.idx = idx
        self.label.configure(text=f"Doc {idx}:")

    def set_enabled(self, enabled=True):
        state = "normal" if enabled else "disabled"
        readonly_state = "readonly" if enabled else "disabled"
        for widget in (self.entry, self.btn_select, self.btn_remove, self.nombre_entry):
            try:
                widget.configure(state=state)
            except Exception:
                pass
        try:
            self.tipo_combo.configure(state=readonly_state)
        except Exception:
            pass



class IoarrDocRow(ttk.Frame):
    def __init__(self, master, idx, remove_callback):
        super().__init__(master)
        self.idx = idx
        self.remove_callback = remove_callback

        self.path_var = tk.StringVar()
        self.tipo_var = tk.StringVar(value="INFORME_TECNICO")

        self.label = ttk.Label(self, text=f"Doc IOARR {idx}:", width=14)
        self.label.grid(row=0, column=0, padx=5, pady=4, sticky="w")

        self.entry = ttk.Entry(self, textvariable=self.path_var, width=64)
        self.entry.grid(row=0, column=1, padx=5, pady=4, sticky="ew")

        self.btn_select = ttk.Button(self, text="Seleccionar", command=self.pick)
        self.btn_select.grid(row=0, column=2, padx=3, pady=4)

        ttk.Label(self, text="Tipo:").grid(row=0, column=3, padx=(12, 3), pady=4, sticky="w")
        self.tipo_combo = ttk.Combobox(
            self,
            textvariable=self.tipo_var,
            values=IOARR_DOC_TYPES,
            state="readonly",
            width=24
        )
        self.tipo_combo.grid(row=0, column=4, padx=3, pady=4, sticky="w")

        self.btn_remove = ttk.Button(self, text="Quitar", command=lambda: self.remove_callback(self))
        self.btn_remove.grid(row=0, column=5, padx=5, pady=4)

        self.columnconfigure(1, weight=1)

        if DND_OK:
            self.entry.drop_target_register(DND_FILES)
            self.entry.dnd_bind("<<Drop>>", self.on_drop)

    def on_drop(self, event):
        value = event.data.strip()
        if value.startswith("{") and value.endswith("}"):
            value = value[1:-1]
        if "} {" in value:
            value = value.split("} {")[0].strip("{}")
        self.path_var.set(os.path.realpath(value))

    def pick(self):
        path = filedialog.askopenfilename(
            filetypes=[
                ("PDF y documentos", "*.pdf *.xlsx *.xls *.doc *.docx *.jpg *.jpeg *.png"),
                ("PDF", "*.pdf"),
                ("Todos", "*.*"),
            ]
        )
        if path:
            self.path_var.set(os.path.realpath(path))

    def get_path(self):
        return self.path_var.get().strip()

    def get_tipo(self):
        return self.tipo_var.get().strip() or "OTRO"

    def set_index(self, idx):
        self.idx = idx
        self.label.configure(text=f"Doc IOARR {idx}:")

    def set_enabled(self, enabled=True):
        state = "normal" if enabled else "disabled"
        readonly_state = "readonly" if enabled else "disabled"
        for widget in (self.entry, self.btn_select, self.btn_remove):
            try:
                widget.configure(state=state)
            except Exception:
                pass
        try:
            self.tipo_combo.configure(state=readonly_state)
        except Exception:
            pass

class LegajosGUIV2(BaseApp):
    def __init__(self):
        super().__init__()
        self.title("LEGAJOS - Interfaz gráfica V2")
        self.geometry("1280x980")
        self.minsize(1120, 860)

        self.current_action = "add"
        self.is_running = False

        self.station_items_all = []
        self.station_items_filtered = []
        self.selected_station_codes = set()

        self.single_station_items_all = []
        self.single_station_items_filtered = []

        self.photo_rows = []
        self.matricula_rows = []
        # Datos específicos para convenios DZ
        self.convenio_frame = ttk.LabelFrame(main, text="Datos del convenio DZ")
        row_convenio = ttk.Frame(self.convenio_frame)
        row_convenio.pack(fill="x", padx=8, pady=6)

        ttk.Label(row_convenio, text="Tipo documento:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.tipo_documento_convenio_var = tk.StringVar(value="CONVENIO")
        self.tipo_documento_convenio_combo = ttk.Combobox(
            row_convenio,
            textvariable=self.tipo_documento_convenio_var,
            values=CONVENIO_DZ_DOC_TYPES,
            state="readonly",
            width=32
        )
        self.tipo_documento_convenio_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")


        self.ioarr_rows = []
        self.temp_generated_file = None
        self.action_buttons = []

        self.build_ui()

    def build_ui(self):
        container = ttk.Frame(self)
        container.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(container, highlightthickness=0)
        self.v_scroll = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.v_scroll.set)

        self.v_scroll.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.main = ttk.Frame(self.canvas, padding=12)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.main, anchor="nw")

        self.main.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        self.bind_all("<MouseWheel>", self._on_mousewheel)
        self.bind_all("<Button-4>", self._on_mousewheel)
        self.bind_all("<Button-5>", self._on_mousewheel)

        ttk.Label(
            self.main,
            text="Gestor gráfico para LEGAJOS_codigo.py - V2",
            font=("Arial", 16, "bold")
        ).pack(anchor="w", pady=(0, 8))

        info = "Arrastra archivos o usa los botones."
        if not DND_OK:
            info += " (Arrastrar/soltar deshabilitado: falta tkinterdnd2)"
        ttk.Label(self.main, text=info).pack(anchor="w", pady=(0, 4))

        self.status_var = tk.StringVar()
        ttk.Label(self.main, textvariable=self.status_var).pack(anchor="w", pady=(0, 8))
        self.update_status_label()

        fuente_frame = ttk.LabelFrame(self.main, text="Archivo fuente")
        fuente_frame.pack(fill="x", pady=8)

        self.src_field = DropField(
            fuente_frame,
            "Archivo fuente:",
            is_dir=False,
            filetypes=[("Todos los archivos", "*.*")]
        )
        self.src_field.pack(fill="x", padx=8, pady=4)

        actions_frame = ttk.LabelFrame(self.main, text="Acciones")
        actions_frame.pack(fill="x", pady=8)

        grid = ttk.Frame(actions_frame)
        grid.pack(fill="x", padx=8, pady=8)

        buttons = [
            ("Agregar documento por estación", "add"),
            ("Agregar mantenimiento grupal sin ruta", "addmantenimiento_grupal"),
            ("Agregar instalación de estación", "addmatricula"),
            ("Agregar checklist de ruta", "addchecklist"),
            ("Agregar estado situacional de ruta", "addestado_situacional"),
            ("Agregar foto de ruta", "addfoto"),
            ("Agregar ruta", "addruta"),
            ("Agregar convenio DZ", "addconvenio_dz"),
            ("Agregar documento IOARR", "addioarr"),
            ("Generar reporte documental anual", "reporte_documental_anual"),
            ("Crear estructura", "init"),
            ("Reconstruir índice", "index"),
            ("Crear ficha DZ", "addficha_dz"),
            ("Crear carpeta ficha DZ", "mk_ficha_dz"),
        ]

        for i, (label, action) in enumerate(buttons):
            btn = ttk.Button(grid, text=label, command=lambda a=action: self.set_action(a))
            btn.grid(row=i // 2, column=i % 2, padx=8, pady=8, sticky="ew")
            self.action_buttons.append(btn)

        grid.columnconfigure(0, weight=1)
        grid.columnconfigure(1, weight=1)

        state_frame = ttk.Frame(self.main)
        state_frame.pack(fill="x", pady=(0, 8))
        ttk.Label(state_frame, text="Acción actual:", font=("Arial", 10, "bold")).pack(side="left")
        self.action_label = ttk.Label(state_frame, text="")
        self.action_label.pack(side="left", padx=8)

        top_exec = ttk.Frame(self.main)
        top_exec.pack(fill="x", pady=(0, 8))

        self.run_button_top = ttk.Button(top_exec, text="Ejecutar acción actual", command=self.run_command)
        self.run_button_top.pack(side="left", padx=5)

        self.clear_button_top = ttk.Button(top_exec, text="Limpiar formulario", command=self.clear_form)
        self.clear_button_top.pack(side="left", padx=5)

        form = ttk.LabelFrame(self.main, text="Datos de la acción")
        form.pack(fill="x", pady=8)

        row1 = ttk.Frame(form)
        row1.pack(fill="x", padx=8, pady=6)

        ttk.Label(row1, text="DZ:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.dz_var = tk.StringVar(value="DZ01")
        dz_values = [f"DZ{str(i).zfill(2)}" for i in range(1, 14)]
        self.dz_combo = ttk.Combobox(row1, textvariable=self.dz_var, values=dz_values, state="readonly", width=10)
        self.dz_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.dz_combo.bind("<<ComboboxSelected>>", self.on_dz_changed)

        ttk.Label(row1, text="Fecha:").grid(row=0, column=2, padx=15, pady=5, sticky="w")
        self.fecha_entry = DateEntry(row1, width=14, date_pattern="dd-mm-y", locale="es_ES")
        self.fecha_entry.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        ttk.Label(row1, text="Código estación:").grid(row=0, column=4, padx=15, pady=5, sticky="w")
        self.codigo_var = tk.StringVar()
        self.codigo_entry = ttk.Entry(row1, textvariable=self.codigo_var, width=18)
        self.codigo_entry.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        row2 = ttk.Frame(form)
        row2.pack(fill="x", padx=8, pady=6)

        ttk.Label(row2, text="Categoría:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.categoria_var = tk.StringVar(value="MANTENIMIENTO")
        categorias = CATEGORIAS_ESTACION
        self.categoria_combo = ttk.Combobox(row2, textvariable=self.categoria_var, values=categorias, state="readonly", width=28)
        self.categoria_combo.bind("<<ComboboxSelected>>", lambda event=None: self.apply_action_visibility())
        self.categoria_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # ttk.Label(row2, text="Ruta:").grid(row=0, column=2, padx=15, pady=5, sticky="w")
        # self.ruta_var = tk.StringVar()
        # self.ruta_entry = ttk.Entry(row2, textvariable=self.ruta_var, width=14)
        # self.ruta_entry.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        ttk.Label(row2, text="Ruta:").grid(row=0, column=2, padx=15, pady=5, sticky="w")
        self.ruta_var = tk.StringVar(value="RUTA_01")
        rutas_values = [f"RUTA_{i:02d}" for i in range(1, 100)]
        self.ruta_entry = ttk.Combobox(
            row2,
            textvariable=self.ruta_var,
            values=rutas_values,
            state="readonly",
            width=12
        )
        self.ruta_entry.grid(row=0, column=3, padx=5, pady=5, sticky="w")


        ttk.Label(row2, text="Tipo ruta:").grid(row=0, column=4, padx=15, pady=5, sticky="w")
        self.tipo_var = tk.StringVar(value="MANTENIMIENTOS")
        self.tipo_combo = ttk.Combobox(
            row2,
            textvariable=self.tipo_var,
            values=["MANTENIMIENTOS", "AFOROS", "INSPECCION"],
            state="readonly",
            width=20
        )
        self.tipo_combo.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        row3 = ttk.Frame(form)
        row3.pack(fill="x", padx=8, pady=6)

        ttk.Label(row3, text="Years (init):").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.years_var = tk.StringVar(value="2024 2025 2026")
        self.years_entry = ttk.Entry(row3, textvariable=self.years_var, width=20)
        self.years_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        ttk.Label(row3, text="Responsable:").grid(row=0, column=2, padx=15, pady=5, sticky="w")
        self.responsable_var = tk.StringVar()
        self.responsable_entry = ttk.Entry(row3, textvariable=self.responsable_var, width=25)
        self.responsable_entry.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        ttk.Label(row3, text="Observaciones:").grid(row=0, column=4, padx=15, pady=5, sticky="w")
        self.obs_var = tk.StringVar()
        self.obs_entry = ttk.Entry(row3, textvariable=self.obs_var, width=28)
        self.obs_entry.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        row4 = ttk.Frame(form)
        row4.pack(fill="x", padx=8, pady=6)

        ttk.Label(row4, text="Year (index):").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.index_year_var = tk.StringVar()
        self.index_year_entry = ttk.Entry(row4, textvariable=self.index_year_var, width=12)
        self.index_year_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        ttk.Label(row4, text="Nombre ficha:").grid(row=0, column=2, padx=15, pady=5, sticky="w")
        self.filename_var = tk.StringVar()
        self.filename_entry = ttk.Entry(row4, textvariable=self.filename_var, width=35)
        self.filename_entry.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        ttk.Label(row4, text="¿Copiar o mover?:").grid(row=0, column=4, padx=15, pady=5, sticky="w")
        self.copy_mode_var = tk.StringVar(value="move")

        copy_mode_frame = ttk.Frame(row4)
        copy_mode_frame.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        ttk.Radiobutton(
            copy_mode_frame,
            text="Sí, copiar",
            variable=self.copy_mode_var,
            value="copy"
        ).pack(side="left", padx=(0, 10))

        ttk.Radiobutton(
            copy_mode_frame,
            text="No, mover",
            variable=self.copy_mode_var,
            value="move"
        ).pack(side="left")

        row5 = ttk.Frame(form)
        row5.pack(fill="x", padx=8, pady=6)

        ttk.Label(row5, text="Sobrescribir ficha:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.overwrite_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(row5, variable=self.overwrite_var).grid(row=0, column=1, padx=5, pady=5, sticky="w")

        ttk.Label(row5, text="Subtipo siniestro:").grid(row=0, column=2, padx=15, pady=5, sticky="w")
        self.siniestro_subtipo_var = tk.StringVar(value="ROBO")
        self.siniestro_subtipo_combo = ttk.Combobox(row5, textvariable=self.siniestro_subtipo_var, values=SINIESTRO_SUBTIPOS, state="readonly", width=30)
        self.siniestro_subtipo_combo.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        ttk.Label(row5, text="Variable/sensor:").grid(row=0, column=4, padx=15, pady=5, sticky="w")
        self.variable_sensor_var = tk.StringVar(value="TEMPERATURA")
        self.variable_sensor_combo = ttk.Combobox(
            row5,
            textvariable=self.variable_sensor_var,
            values=SENSOR_VARIABLES,
            state="readonly",
            width=24
        )
        self.variable_sensor_combo.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        row6 = ttk.Frame(form)
        row6.pack(fill="x", padx=8, pady=6)

        ttk.Label(row6, text="Nombre IOARR:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.nombre_ioarr_var = tk.StringVar(value="")
        self.nombre_ioarr_entry = ttk.Entry(row6, textvariable=self.nombre_ioarr_var, width=38)
        self.nombre_ioarr_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        self.tipo_doc_ioarr_label = ttk.Label(row6, text="Tipo doc IOARR:")
        self.tipo_doc_ioarr_var = tk.StringVar(value="INFORME_TECNICO")
        self.tipo_doc_ioarr_combo = ttk.Combobox(
            row6,
            textvariable=self.tipo_doc_ioarr_var,
            values=IOARR_DOC_TYPES,
            state="readonly",
            width=24
        )
        # Este selector general queda oculto; el tipo documental ahora se elige por cada documento IOARR.
        # self.tipo_doc_ioarr_combo.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        self.single_station_frame = ttk.LabelFrame(self.main, text="Buscar estación para acciones individuales")
        self.single_station_frame.pack(fill="x", pady=8)

        single_top = ttk.Frame(self.single_station_frame)
        single_top.pack(fill="x", padx=8, pady=6)

        ttk.Label(single_top, text="Buscar estación:").pack(side="left", padx=(0, 5))
        self.search_single_station_var = tk.StringVar()
        self.search_single_station_var.trace_add("write", lambda *args: self.filter_single_station_list())
        self.search_single_station_entry = ttk.Entry(single_top, textvariable=self.search_single_station_var, width=40)
        self.search_single_station_entry.pack(side="left", padx=5)

        ttk.Button(single_top, text="Cargar estaciones DZ", command=self.load_single_station_candidates).pack(side="left", padx=8)
        ttk.Button(single_top, text="Limpiar código", command=self.clear_single_station_code).pack(side="left", padx=5)
        ttk.Button(single_top, text="Desbloquear código", command=self.unlock_codigo_entry).pack(side="left", padx=5)

        single_list_frame = ttk.Frame(self.single_station_frame)
        single_list_frame.pack(fill="both", expand=True, padx=8, pady=6)

        self.single_station_listbox = tk.Listbox(single_list_frame, height=6)
        self.single_station_listbox.pack(side="left", fill="both", expand=True)
        self.single_station_listbox.bind("<<ListboxSelect>>", self.select_single_station_from_list)

        single_scroll = ttk.Scrollbar(single_list_frame, orient="vertical", command=self.single_station_listbox.yview)
        single_scroll.pack(side="right", fill="y")
        self.single_station_listbox.configure(yscrollcommand=single_scroll.set)

        self.selected_single_station_label = ttk.Label(
            self.single_station_frame,
            text="Estación seleccionada: Ninguna",
            font=("Arial", 10, "bold")
        )
        self.selected_single_station_label.pack(anchor="w", padx=8, pady=(0, 8))

        self.matricula_frame = ttk.LabelFrame(self.main, text="Documentos de instalación de estación")
        self.matricula_frame.pack(fill="x", pady=8)

        matricula_top = ttk.Frame(self.matricula_frame)
        matricula_top.pack(fill="x", padx=8, pady=6)

        ttk.Button(matricula_top, text="Agregar documento", command=self.add_matricula_row).pack(side="left", padx=5)
        ttk.Button(matricula_top, text="Quitar último", command=self.remove_last_matricula_row).pack(side="left", padx=5)
        ttk.Label(
            matricula_top,
            text="Cada fila permite elegir un archivo, su tipo técnico y el nombre final editable."
        ).pack(side="left", padx=10)

        self.matricula_rows_container = ttk.Frame(self.matricula_frame)
        self.matricula_rows_container.pack(fill="x", padx=8, pady=6)

        for _ in range(3):
            self.add_matricula_row()

        self.ioarr_frame = ttk.LabelFrame(self.main, text="Documentos IOARR")
        self.ioarr_frame.pack(fill="x", pady=8)

        ioarr_top = ttk.Frame(self.ioarr_frame)
        ioarr_top.pack(fill="x", padx=8, pady=6)

        ttk.Button(ioarr_top, text="Agregar documento IOARR", command=self.add_ioarr_row).pack(side="left", padx=5)
        ttk.Button(ioarr_top, text="Quitar último", command=self.remove_last_ioarr_row).pack(side="left", padx=5)
        ttk.Label(
            ioarr_top,
            text="Carga uno o varios documentos IOARR y asigna el tipo documental a cada archivo."
        ).pack(side="left", padx=10)

        self.ioarr_rows_container = ttk.Frame(self.ioarr_frame)
        self.ioarr_rows_container.pack(fill="x", padx=8, pady=6)

        for _ in range(4):
            self.add_ioarr_row()

        self.photos_frame = ttk.LabelFrame(self.main, text="Fotos para unir en orden (solo para Agregar foto)")
        self.photos_frame.pack(fill="x", pady=8)

        photos_top = ttk.Frame(self.photos_frame)
        photos_top.pack(fill="x", padx=8, pady=6)

        ttk.Button(photos_top, text="Agregar casilla de foto", command=self.add_photo_row).pack(side="left", padx=5)
        ttk.Button(photos_top, text="Quitar última", command=self.remove_last_photo_row).pack(side="left", padx=5)
        ttk.Label(
            photos_top,
            text="Puedes cargar imágenes o PDF. También puedes arrastrarlos a cada casilla. Todo se unirá en un solo PDF respetando el orden. Usa ↑ ↓ para reordenar."
        ).pack(side="left", padx=10)

        self.photos_rows_container = ttk.Frame(self.photos_frame)
        self.photos_rows_container.pack(fill="x", padx=8, pady=6)

        self.pdf_preview_var = tk.StringVar(value="PDF a generar: fotos_unidas.pdf | Cantidad de archivos: 0")
        ttk.Label(self.photos_frame, textvariable=self.pdf_preview_var).pack(anchor="w", padx=8, pady=(0, 8))

        for _ in range(3):
            self.add_photo_row()

        stations_frame = ttk.LabelFrame(self.main, text="Estaciones de la DZ (selección múltiple)")
        stations_frame.pack(fill="both", expand=False, pady=8)

        top_st = ttk.Frame(stations_frame)
        top_st.pack(fill="x", padx=8, pady=6)

        ttk.Label(top_st, text="Buscar:").pack(side="left", padx=(0, 5))
        self.search_station_var = tk.StringVar()
        self.search_station_var.trace_add("write", lambda *args: self.filter_station_list())
        self.search_station_entry = ttk.Entry(top_st, textvariable=self.search_station_var, width=35)
        self.search_station_entry.pack(side="left", padx=5)

        ttk.Button(top_st, text="Cargar estaciones", command=self.load_stations_for_dz).pack(side="left", padx=8)
        ttk.Button(top_st, text="Seleccionar todas", command=self.select_all_stations).pack(side="left", padx=5)
        ttk.Button(top_st, text="Limpiar selección", command=self.clear_station_selection).pack(side="left", padx=5)
        ttk.Button(top_st, text="Ver seleccionadas", command=self.show_selected_stations).pack(side="left", padx=5)

        list_frame = ttk.Frame(stations_frame)
        list_frame.pack(fill="both", expand=True, padx=8, pady=6)

        self.station_listbox = tk.Listbox(list_frame, selectmode=tk.MULTIPLE, height=12)
        self.station_listbox.pack(side="left", fill="both", expand=True)
        self.station_listbox.bind("<<ListboxSelect>>", self.sync_visible_selection)

        station_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.station_listbox.yview)
        station_scroll.pack(side="right", fill="y")
        self.station_listbox.configure(yscrollcommand=station_scroll.set)

        help_frame = ttk.Frame(stations_frame)
        help_frame.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Label(
            help_frame,
            text="Usa Ctrl o Shift para seleccionar varias estaciones. Se usa en: Crear estructura, Agregar ruta, Agregar mantenimiento grupal, Agregar convenio DZ y Agregar documento IOARR."
        ).pack(anchor="w")

        exec_frame = ttk.Frame(self.main)
        exec_frame.pack(fill="x", pady=10)

        self.run_button_bottom = ttk.Button(exec_frame, text="Ejecutar acción actual", command=self.run_command)
        self.run_button_bottom.pack(side="left", padx=5)

        self.clear_button_bottom = ttk.Button(exec_frame, text="Limpiar formulario", command=self.clear_form)
        self.clear_button_bottom.pack(side="left", padx=5)

        ttk.Button(exec_frame, text="Verificar archivos base", command=self.verify_base_files).pack(side="left", padx=5)

        out = ttk.LabelFrame(self.main, text="Salida")
        out.pack(fill="both", expand=True, pady=8)

        self.console = tk.Text(out, wrap="word", height=18)
        self.console.pack(side="left", fill="both", expand=True, padx=8, pady=8)

        scrollbar = ttk.Scrollbar(out, orient="vertical", command=self.console.yview)
        scrollbar.pack(side="right", fill="y", pady=8)
        self.console.configure(yscrollcommand=scrollbar.set)

        self.log("Interfaz lista.\n")
        self.set_action("add")
        self.load_stations_for_dz()
        self.load_single_station_candidates()
        self.update_pdf_preview_label()

    def _on_frame_configure(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self.canvas_window, width=event.width)

    def _on_mousewheel(self, event):
        if hasattr(event, "delta") and event.delta:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        elif getattr(event, "num", None) == 4:
            self.canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5:
            self.canvas.yview_scroll(1, "units")

    def update_pdf_preview_label(self):
        count = len(self.get_photo_paths())
        self.pdf_preview_var.set(f"PDF a generar: fotos_unidas.pdf | Cantidad de archivos: {count}")

    def add_matricula_row(self):
        row = MatriculaDocRow(
            self.matricula_rows_container,
            len(self.matricula_rows) + 1,
            self.remove_matricula_row,
            self.refresh_matricula_final_names,
        )
        row.pack(fill="x", pady=2)
        self.matricula_rows.append(row)
        self.refresh_matricula_final_names()

    def remove_matricula_row(self, row):
        if len(self.matricula_rows) <= 1:
            messagebox.showinfo("Matrículas", "Debe quedar al menos una fila de documento.")
            return
        row.destroy()
        self.matricula_rows.remove(row)
        self.rebuild_matricula_rows()
        self.refresh_matricula_final_names()

    def remove_last_matricula_row(self):
        if self.matricula_rows:
            self.remove_matricula_row(self.matricula_rows[-1])

    def rebuild_matricula_rows(self):
        for widget in self.matricula_rows_container.winfo_children():
            widget.pack_forget()
        for i, row in enumerate(self.matricula_rows, start=1):
            row.set_index(i)
            row.pack(fill="x", pady=2)

    def get_matricula_docs(self):
        docs = []
        for row in self.matricula_rows:
            path = row.get_path()
            if path:
                docs.append((path, row.get_tipo(), row.get_nombre_final()))
        return docs

    def get_selected_station_folder_hint(self):
        codigo = self.codigo_var.get().strip()
        if not codigo:
            return "ESTACION"
        for code, texto in self.single_station_items_all:
            if code == codigo:
                parts = [p.strip() for p in texto.split("|")]
                if len(parts) >= 3:
                    return slug_text(f"{parts[0]}-{parts[1]}-{parts[2]}")
        return slug_text(codigo)

    def refresh_matricula_final_names(self):
        if not hasattr(self, "matricula_rows"):
            return
        dz = self.dz_var.get().strip().upper() or "DZ"
        fecha = self.fecha_entry.get().strip() if hasattr(self, "fecha_entry") else "FECHA"
        try:
            # DateEntry usa dd-mm-y; convertimos solo para el nombre.
            d, m, y = fecha.split("-")
            fecha_name = f"20{y}-{m}-{d}" if len(y) == 2 else f"{y}-{m}-{d}"
        except Exception:
            fecha_name = slug_text(fecha)
        estacion = self.get_selected_station_folder_hint()

        counts = {}
        for row in self.matricula_rows:
            if not row.get_path():
                continue
            tipo = row.get_tipo() or "OTRO"
            ext = os.path.splitext(row.get_path())[1].lower() or ".pdf"
            base = f"{tipo}_{dz}_{estacion}_{fecha_name}"
            counts[base] = counts.get(base, 0) + 1
            suffix = f"_{counts[base]:02d}" if counts[base] > 1 else ""
            suggested = f"{base}{suffix}{ext}"
            # Solo autocompleta si está vacío o si parece un nombre autogenerado anterior.
            current = row.nombre_final_var.get().strip()
            current_upper = current.upper()
            if not current or any(current_upper.startswith(t + "_") for t in MATRICULA_DOC_TYPES):
                row.nombre_final_var.set(suggested)

    def update_status_label(self):
        script_ok = os.path.isfile(SCRIPT_DEFAULT)
        maestra_ok = os.path.isfile(MAESTRA_DEFAULT)
        self.status_var.set(
            f"Script detectado: {'OK' if script_ok else 'NO'} | "
            f"Excel maestro detectado: {'OK' if maestra_ok else 'NO'} | "
            f"Carpeta base: {BASE_DIR}"
        )

    def verify_base_files(self):
        self.update_status_label()
        msg = (
            f"LEGAJOS_codigo.py:\n{SCRIPT_DEFAULT}\n"
            f"{'Encontrado' if os.path.isfile(SCRIPT_DEFAULT) else 'No encontrado'}\n\n"
            f"MaestraEstaciones.xlsx:\n{MAESTRA_DEFAULT}\n"
            f"{'Encontrado' if os.path.isfile(MAESTRA_DEFAULT) else 'No encontrado'}"
        )
        messagebox.showinfo("Verificación", msg)

    def log(self, text):
        self.console.insert("end", text)
        self.console.see("end")
        self.update_idletasks()

    def norm_header(self, text):
        text = str(text).lower().strip()
        replacements = {"á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n"}
        for a, b in replacements.items():
            text = text.replace(a, b)
        return text

    def find_header_index(self, headers_norm, posibles):
        posibles_norm = [self.norm_header(p) for p in posibles]
        for i, h in enumerate(headers_norm):
            if h in posibles_norm:
                return i
        return None

    def safe_cell(self, row, idx):
        if idx is None or idx >= len(row):
            return ""
        value = row[idx]
        return str(value).strip() if value is not None else ""

    def normalize_dz_value(self, value):
        value = str(value).strip().upper()
        if not value:
            return ""
        if value.startswith("DZ"):
            num = value.replace("DZ", "").strip()
            if num.isdigit():
                return f"DZ{int(num):02d}"
            return value
        if value.isdigit():
            return f"DZ{int(value):02d}"
        return value

    def set_enabled(self, widget, enabled=True):
        state = "normal" if enabled else "disabled"
        try:
            widget.configure(state=state)
        except Exception:
            pass

    def set_running_state(self, running: bool):
        self.is_running = running
        run_state = "disabled" if running else "normal"
        clear_state = "disabled" if running else "normal"

        self.run_button_top.configure(state=run_state)
        self.run_button_bottom.configure(state=run_state)
        self.clear_button_top.configure(state=clear_state)
        self.clear_button_bottom.configure(state=clear_state)

        for btn in self.action_buttons:
            btn.configure(state="disabled" if running else "normal")

    def add_ioarr_row(self):
        row = IoarrDocRow(
            self.ioarr_rows_container,
            len(self.ioarr_rows) + 1,
            self.remove_ioarr_row,
        )
        row.pack(fill="x", pady=2)
        self.ioarr_rows.append(row)

    def remove_ioarr_row(self, row):
        if len(self.ioarr_rows) <= 1:
            messagebox.showinfo("IOARR", "Debe quedar al menos una fila de documento IOARR.")
            return
        row.destroy()
        self.ioarr_rows.remove(row)
        self.rebuild_ioarr_rows()

    def remove_last_ioarr_row(self):
        if self.ioarr_rows:
            self.remove_ioarr_row(self.ioarr_rows[-1])

    def rebuild_ioarr_rows(self):
        for widget in self.ioarr_rows_container.winfo_children():
            widget.pack_forget()
        for i, row in enumerate(self.ioarr_rows, start=1):
            row.set_index(i)
            row.pack(fill="x", pady=2)

    def get_ioarr_docs(self):
        docs = []
        for row in self.ioarr_rows:
            path = row.get_path()
            if path:
                docs.append((path, row.get_tipo()))
        return docs

    def add_photo_row(self):
        row = MultiPhotoRow(
            self.photos_rows_container,
            len(self.photo_rows) + 1,
            self.remove_photo_row,
            self.move_photo_row_up,
            self.move_photo_row_down,
            self.update_pdf_preview_label,
        )
        row.pack(fill="x", pady=2)
        self.photo_rows.append(row)
        self.update_pdf_preview_label()

    def remove_photo_row(self, row):
        if len(self.photo_rows) <= 1:
            messagebox.showinfo("Fotos", "Debe quedar al menos una casilla de foto.")
            return
        row.destroy()
        self.photo_rows.remove(row)
        self.rebuild_photo_rows()
        self.update_pdf_preview_label()

    def remove_last_photo_row(self):
        if self.photo_rows:
            self.remove_photo_row(self.photo_rows[-1])

    def move_photo_row_up(self, row):
        idx = self.photo_rows.index(row)
        if idx == 0:
            return
        self.photo_rows[idx], self.photo_rows[idx - 1] = self.photo_rows[idx - 1], self.photo_rows[idx]
        self.rebuild_photo_rows()
        self.update_pdf_preview_label()

    def move_photo_row_down(self, row):
        idx = self.photo_rows.index(row)
        if idx == len(self.photo_rows) - 1:
            return
        self.photo_rows[idx], self.photo_rows[idx + 1] = self.photo_rows[idx + 1], self.photo_rows[idx]
        self.rebuild_photo_rows()
        self.update_pdf_preview_label()

    def rebuild_photo_rows(self):
        for widget in self.photos_rows_container.winfo_children():
            widget.pack_forget()
        for i, row in enumerate(self.photo_rows, start=1):
            row.set_index(i)
            row.pack(fill="x", pady=2)

    def get_photo_paths(self):
        return [r.get() for r in self.photo_rows if r.get()]

    def build_photos_pdf(self):
        photo_paths = self.get_photo_paths()
        if not photo_paths:
            raise ValueError("Debes cargar al menos un archivo.")

        temp_dir = tempfile.mkdtemp(prefix="legajos_fotos_")
        final_pdf_path = os.path.join(temp_dir, "fotos_unidas.pdf")

        writer = PdfWriter()

        for idx, path in enumerate(photo_paths, start=1):
            if not os.path.isfile(path):
                raise ValueError(f"No existe el archivo: {path}")

            ext = os.path.splitext(path)[1].lower()

            if ext == ".pdf":
                writer.append(path)
            else:
                img = Image.open(path).convert("RGB")
                temp_img_pdf = os.path.join(temp_dir, f"img_{idx}.pdf")
                img.save(temp_img_pdf, "PDF")
                writer.append(temp_img_pdf)

        with open(final_pdf_path, "wb") as f:
            writer.write(f)

        self.temp_generated_file = final_pdf_path
        return final_pdf_path

    def set_action(self, action):
        self.current_action = action
        names = {
            "add": "Agregar documento por estación",
            "addmantenimiento_grupal": "Agregar mantenimiento grupal sin ruta",
            "addmatricula": "Agregar instalación de estación",
            "addchecklist": "Agregar checklist de ruta",
            "addestado_situacional": "Agregar estado situacional de ruta",
            "addfoto": "Agregar foto de ruta",
            "addruta": "Agregar ruta",
            "addconvenio_dz": "Agregar convenio DZ",
            "addioarr": "Agregar documento IOARR",
            "reporte_documental_anual": "Generar reporte documental anual",
            "init": "Crear estructura",
            "index": "Reconstruir índice",
            "addficha_dz": "Crear ficha DZ",
            "mk_ficha_dz": "Crear carpeta ficha DZ",
        }
        self.action_label.config(text=names.get(action, action))
        self.apply_action_visibility()

    def apply_action_visibility(self):
        action = self.current_action

        src_enabled = action in {"add", "addmantenimiento_grupal", "addchecklist", "addestado_situacional", "addruta", "addconvenio_dz"}
        self.set_enabled(self.src_field.entry, src_enabled)
        self.set_enabled(self.src_field.button, src_enabled)

        self.set_enabled(self.fecha_entry, action in {"add", "addmatricula", "addmantenimiento_grupal", "addchecklist", "addestado_situacional", "addfoto", "addruta", "addconvenio_dz", "addioarr"})
        self.set_enabled(self.codigo_entry, action in {"add", "addmatricula", "addchecklist", "addestado_situacional", "addfoto", "index"})
        self.set_enabled(self.categoria_combo, action == "add")
        self.set_enabled(self.ruta_entry, action in {"addruta", "addchecklist", "addestado_situacional", "addfoto"})
        self.set_enabled(self.tipo_combo, action == "addruta")
        self.set_enabled(self.years_entry, action == "init")
        self.set_enabled(self.responsable_entry, action in {"addruta", "addmantenimiento_grupal"})
        self.set_enabled(self.obs_entry, action in {"addruta", "addmantenimiento_grupal", "addconvenio_dz", "addioarr"})
        self.set_enabled(self.index_year_entry, action in {"index", "reporte_documental_anual"})
        self.set_enabled(self.filename_entry, action == "addficha_dz")
        self.set_enabled(self.siniestro_subtipo_combo, action == "add" and self.categoria_var.get() == "SINIESTROS")
        self.set_enabled(self.variable_sensor_combo, action == "add" and self.categoria_var.get() == "CERTIFICADO_COMPROBACION_SENSOR")
        self.set_enabled(self.nombre_ioarr_entry, action == "addioarr")
        self.set_enabled(self.tipo_doc_ioarr_combo, False)

        stations_enabled = action in {"init", "addruta", "addmantenimiento_grupal", "addconvenio_dz", "addioarr"}
        self.set_enabled(self.station_listbox, stations_enabled)
        self.set_enabled(self.search_station_entry, stations_enabled)

        single_station_enabled = action in {"add", "addmatricula", "addchecklist", "addestado_situacional", "addfoto", "index"}
        self.set_enabled(self.search_single_station_entry, single_station_enabled)
        self.set_enabled(self.single_station_listbox, single_station_enabled)

        matricula_enabled = action == "addmatricula"
        if hasattr(self, "matricula_frame"):
            if matricula_enabled:
                self.matricula_frame.pack(fill="x", pady=8, after=self.selected_single_station_label.master)
                self.refresh_matricula_final_names()
            else:
                self.matricula_frame.pack_forget()
        for row in getattr(self, "matricula_rows", []):
            row.set_enabled(matricula_enabled)

        ioarr_enabled = action == "addioarr"
        convenio_enabled = action == "addconvenio_dz"
        if hasattr(self, "ioarr_frame"):
            if ioarr_enabled:
                # Se coloca en una posición fija: debajo del buscador de estación individual.
                # No se usa matrícula como referencia porque esa sección se oculta en IOARR.
                self.ioarr_frame.pack(fill="x", pady=8, after=self.single_station_frame)
            else:
                self.ioarr_frame.pack_forget()
        for row in getattr(self, "ioarr_rows", []):
            row.set_enabled(ioarr_enabled)

        photos_enabled = action == "addfoto"
        for row in self.photo_rows:
            self.set_enabled(row.entry, photos_enabled)
            self.set_enabled(row.btn_select, photos_enabled)
            self.set_enabled(row.btn_remove, photos_enabled)
            self.set_enabled(row.btn_up, photos_enabled)
            self.set_enabled(row.btn_down, photos_enabled)

        if self.codigo_var.get().strip() and action in {"add", "addmatricula", "addchecklist", "addestado_situacional", "addfoto", "index"}:
            try:
                self.codigo_entry.configure(state="disabled")
            except Exception:
                pass

    def load_single_station_candidates(self):
        dz_selected = self.dz_var.get().strip().upper()

        self.single_station_listbox.delete(0, "end")
        self.single_station_items_all = []
        self.single_station_items_filtered = []

        if not os.path.isfile(MAESTRA_DEFAULT):
            self.log(f"⚠️ No se encontró el Excel maestro en: {MAESTRA_DEFAULT}\n")
            return

        try:
            wb = load_workbook(MAESTRA_DEFAULT, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                self.log("⚠️ El Excel maestro está vacío.\n")
                return

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            headers_norm = [self.norm_header(h) for h in headers]

            idx_codigo = self.find_header_index(headers_norm, ["codigo", "código"])
            idx_nombre = self.find_header_index(headers_norm, ["nombre"])
            idx_clas = self.find_header_index(headers_norm, ["clasificacion", "clasificación"])
            idx_dz = self.find_header_index(headers_norm, ["dz", "direccion zonal", "dirección zonal"])

            if None in (idx_codigo, idx_nombre, idx_clas, idx_dz):
                self.log("⚠️ El Excel debe tener columnas como Código, Nombre, Clasificación y DZ.\n")
                return

            items = []
            for row in rows[1:]:
                codigo = self.safe_cell(row, idx_codigo)
                nombre = self.safe_cell(row, idx_nombre)
                clas = self.safe_cell(row, idx_clas)
                dz = self.normalize_dz_value(self.safe_cell(row, idx_dz))

                if codigo and dz == dz_selected:
                    items.append((codigo, f"{codigo} | {clas} | {nombre}"))

            items.sort(key=lambda x: x[1])
            self.single_station_items_all = items
            self.single_station_items_filtered = items.copy()

            for _, texto in items:
                self.single_station_listbox.insert("end", texto)

            self.log(f"✅ Se cargaron {len(items)} estaciones individuales para {dz_selected}.\n")

        except Exception as e:
            self.log(f"❌ Error al cargar estaciones individuales: {e}\n")

    def filter_single_station_list(self):
        filtro = self.search_single_station_var.get().strip().lower()
        self.single_station_listbox.delete(0, "end")

        if not filtro:
            self.single_station_items_filtered = self.single_station_items_all.copy()
        else:
            self.single_station_items_filtered = [
                item for item in self.single_station_items_all
                if filtro in item[0].lower() or filtro in item[1].lower()
            ]

        for _, texto in self.single_station_items_filtered:
            self.single_station_listbox.insert("end", texto)

    def select_single_station_from_list(self, event=None):
        selected = self.single_station_listbox.curselection()
        if not selected:
            return

        idx = selected[0]
        codigo, texto = self.single_station_items_filtered[idx]

        self.codigo_entry.configure(state="normal")
        self.codigo_var.set(codigo)
        self.codigo_entry.configure(state="disabled")
        self.selected_single_station_label.config(text=f"Estación seleccionada: {texto}")
        self.refresh_matricula_final_names()

    def clear_single_station_code(self):
        self.codigo_entry.configure(state="normal")
        self.codigo_var.set("")
        self.selected_single_station_label.config(text="Estación seleccionada: Ninguna")
        self.refresh_matricula_final_names()

    def unlock_codigo_entry(self):
        self.codigo_entry.configure(state="normal")

    def on_dz_changed(self, event=None):
        self.search_station_var.set("")
        self.search_single_station_var.set("")
        self.load_stations_for_dz()
        self.load_single_station_candidates()
        self.clear_single_station_code()
        self.refresh_matricula_final_names()

    def load_stations_for_dz(self):
        dz_selected = self.dz_var.get().strip().upper()
        self.station_listbox.delete(0, "end")
        self.station_items_all = []
        self.station_items_filtered = []
        self.selected_station_codes = set()

        if not os.path.isfile(MAESTRA_DEFAULT):
            self.log(f"⚠️ No se encontró el Excel maestro en: {MAESTRA_DEFAULT}\n")
            return

        try:
            wb = load_workbook(MAESTRA_DEFAULT, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                self.log("⚠️ El Excel maestro está vacío.\n")
                return

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            headers_norm = [self.norm_header(h) for h in headers]

            idx_codigo = self.find_header_index(headers_norm, ["codigo", "código"])
            idx_nombre = self.find_header_index(headers_norm, ["nombre"])
            idx_clas = self.find_header_index(headers_norm, ["clasificacion", "clasificación"])
            idx_dz = self.find_header_index(headers_norm, ["dz", "direccion zonal", "dirección zonal"])

            if None in (idx_codigo, idx_nombre, idx_clas, idx_dz):
                self.log("⚠️ El Excel debe tener columnas como Código, Nombre, Clasificación y DZ.\n")
                return

            items = []
            for row in rows[1:]:
                codigo = self.safe_cell(row, idx_codigo)
                nombre = self.safe_cell(row, idx_nombre)
                clas = self.safe_cell(row, idx_clas)
                dz = self.normalize_dz_value(self.safe_cell(row, idx_dz))

                if codigo and dz == dz_selected:
                    items.append((codigo, f"{codigo} | {clas} | {nombre}"))

            items.sort(key=lambda x: x[1])
            self.station_items_all = items
            self.station_items_filtered = items.copy()

            for _, texto in items:
                self.station_listbox.insert("end", texto)

            self.log(f"✅ Se cargaron {len(items)} estaciones para {dz_selected}.\n")

        except Exception as e:
            self.log(f"❌ Error al leer estaciones del Excel: {e}\n")

    def filter_station_list(self):
        filtro = self.search_station_var.get().strip().lower()
        self.station_listbox.delete(0, "end")

        if not filtro:
            self.station_items_filtered = self.station_items_all.copy()
        else:
            self.station_items_filtered = [
                item for item in self.station_items_all
                if filtro in item[0].lower() or filtro in item[1].lower()
            ]

        for idx, (codigo, texto) in enumerate(self.station_items_filtered):
            self.station_listbox.insert("end", texto)
            if codigo in self.selected_station_codes:
                self.station_listbox.select_set(idx)

    def sync_visible_selection(self, event=None):
        visible_codes = [codigo for codigo, _ in self.station_items_filtered]
        selected_indexes = set(self.station_listbox.curselection())

        for idx, codigo in enumerate(visible_codes):
            if idx in selected_indexes:
                self.selected_station_codes.add(codigo)
            else:
                self.selected_station_codes.discard(codigo)

        if self.current_action == "addioarr":
            self.update_nombre_ioarr_from_selected_stations()

    def select_all_stations(self):
        for codigo, _ in self.station_items_filtered:
            self.selected_station_codes.add(codigo)
        self.filter_station_list()

    def update_nombre_ioarr_from_selected_stations(self):
        """
        Llena automáticamente el campo Nombre IOARR usando los nombres de las
        estaciones seleccionadas en la lista múltiple.

        Ejemplo:
        PUENTE_BRENA-TULUMAYO
        """
        if not hasattr(self, "nombre_ioarr_var"):
            return

        codigos = self.get_selected_station_codes()
        if not codigos:
            return

        nombres = []
        for code in codigos:
            texto = self.get_station_text_by_code(code)
            parts = [p.strip() for p in texto.split("|")]
            if len(parts) >= 3:
                nombre = parts[2]
            else:
                nombre = texto

            nombre_slug = slug_text(nombre)
            if nombre_slug and nombre_slug not in nombres:
                nombres.append(nombre_slug)

        if nombres:
            self.nombre_ioarr_var.set("-".join(nombres))

    def clear_station_selection(self):
        for codigo, _ in self.station_items_filtered:
            self.selected_station_codes.discard(codigo)
        self.filter_station_list()

    def get_selected_station_codes(self):
        return sorted(self.selected_station_codes)

    def get_station_text_by_code(self, code):
        for codigo, texto in self.station_items_all:
            if codigo == code:
                return texto
        return code

    def show_selected_stations(self):
        codigos = self.get_selected_station_codes()
        if not codigos:
            messagebox.showinfo("Estaciones seleccionadas", "No has seleccionado ninguna estación.")
            return

        win = tk.Toplevel(self)
        win.title("Estaciones seleccionadas")
        win.geometry("700x420")
        win.transient(self)
        win.grab_set()

        top = ttk.Frame(win, padding=10)
        top.pack(fill="both", expand=True)

        ttk.Label(
            top,
            text=f"Cantidad de estaciones seleccionadas: {len(codigos)}",
            font=("Arial", 10, "bold")
        ).pack(anchor="w", pady=(0, 8))

        list_frame = ttk.Frame(top)
        list_frame.pack(fill="both", expand=True)

        lb = tk.Listbox(list_frame, selectmode=tk.MULTIPLE)
        lb.pack(side="left", fill="both", expand=True)

        scroll = ttk.Scrollbar(list_frame, orient="vertical", command=lb.yview)
        scroll.pack(side="right", fill="y")
        lb.configure(yscrollcommand=scroll.set)

        items = []
        for code in codigos:
            texto = self.get_station_text_by_code(code)
            items.append((code, texto))
            lb.insert("end", texto)

        btns = ttk.Frame(top)
        btns.pack(fill="x", pady=10)

        def quitar_seleccionadas():
            selected_idx = list(lb.curselection())
            if not selected_idx:
                messagebox.showinfo("Quitar", "Selecciona una o más estaciones para quitarlas.", parent=win)
                return

            to_remove = [items[i][0] for i in selected_idx]
            for code in to_remove:
                self.selected_station_codes.discard(code)

            self.filter_station_list()
            win.destroy()

        def limpiar_todas():
            if not messagebox.askyesno("Confirmar", "¿Deseas quitar todas las estaciones seleccionadas?", parent=win):
                return
            self.selected_station_codes.clear()
            self.filter_station_list()
            win.destroy()

        ttk.Button(btns, text="Quitar seleccionadas", command=quitar_seleccionadas).pack(side="left", padx=5)
        ttk.Button(btns, text="Limpiar todas", command=limpiar_todas).pack(side="left", padx=5)
        ttk.Button(btns, text="Cerrar", command=win.destroy).pack(side="right", padx=5)

    def clear_after_success(self):
        """
        Limpia campos sensibles después de ejecutar correctamente una acción,
        para evitar reutilizar estaciones o archivos por error.

        Mantiene:
        - acción actual
        - DZ
        - fecha
        - ruta

        Limpia:
        - archivo fuente
        - estaciones seleccionadas
        - fotos/documentos múltiples
        - nombre IOARR
        """
        # Limpiar archivo fuente
        if hasattr(self, "src_var"):
            self.src_var.set("")

        # Limpiar estación individual
        if hasattr(self, "station_search_var"):
            self.station_search_var.set("")
        if hasattr(self, "station_code_var"):
            self.station_code_var.set("")

        # Limpiar selección múltiple de estaciones
        if hasattr(self, "selected_station_codes"):
            self.selected_station_codes.clear()
        if hasattr(self, "station_listbox"):
            try:
                self.station_listbox.selection_clear(0, tk.END)
            except Exception:
                pass

        # Refrescar lista visible si existe
        if hasattr(self, "filter_station_list"):
            try:
                self.filter_station_list()
            except Exception:
                pass

        # Limpiar fotos si existen
        if hasattr(self, "photo_rows"):
            self.photo_rows.clear()
            if hasattr(self, "rebuild_photo_rows"):
                self.rebuild_photo_rows()

        # Limpiar documentos IOARR si existen
        if hasattr(self, "ioarr_rows"):
            self.ioarr_rows.clear()
            if hasattr(self, "rebuild_ioarr_rows"):
                self.rebuild_ioarr_rows()
            if hasattr(self, "add_ioarr_row"):
                self.add_ioarr_row()

        # Limpiar nombre IOARR autogenerado
        if hasattr(self, "nombre_ioarr_var"):
            self.nombre_ioarr_var.set("")

        # Limpiar nombre descriptivo del convenio
        if hasattr(self, "nombre_convenio_var"):
            self.nombre_convenio_var.set("")

        # Mantener variable/sensor en un valor seguro por defecto
        if hasattr(self, "variable_sensor_var"):
            self.variable_sensor_var.set("TEMPERATURA")

    def clear_form(self):
        if self.is_running:
            return

        self.codigo_entry.configure(state="normal")
        self.src_field.set("")
        self.fecha_entry.set_date(date.today())
        self.codigo_var.set("")
        self.ruta_var.set("RUTA_01")
        self.responsable_var.set("")
        self.obs_var.set("")
        self.index_year_var.set("")
        self.filename_var.set("")
        self.search_station_var.set("")
        self.search_single_station_var.set("")
        self.copy_mode_var.set("move")
        self.overwrite_var.set(False)
        self.siniestro_subtipo_var.set("ROBO")
        self.variable_sensor_var.set("TEMPERATURA")
        self.nombre_ioarr_var.set("")
        self.tipo_doc_ioarr_var.set("INFORME_TECNICO")

        self.station_listbox.delete(0, "end")
        self.station_items_all = []
        self.station_items_filtered = []
        self.selected_station_codes = set()

        self.single_station_listbox.delete(0, "end")
        self.single_station_items_all = []
        self.single_station_items_filtered = []
        self.selected_single_station_label.config(text="Estación seleccionada: Ninguna")

        for row in self.photo_rows:
            row.var.set("")
            row.update_preview()

        for row in self.matricula_rows:
            row.path_var.set("")
            row.tipo_var.set("ACTA_INSTALACION")
            row.nombre_final_var.set("")

        self.console.delete("1.0", "end")
        self.log("Formulario limpiado.\n")
        self.update_status_label()
        self.update_pdf_preview_label()

    def validate(self):
        action = self.current_action

        if not os.path.isfile(SCRIPT_DEFAULT):
            return False, f"No se encontró LEGAJOS_codigo.py en:\n{SCRIPT_DEFAULT}"

        if action in {"add", "addmatricula", "addmantenimiento_grupal", "addchecklist", "addestado_situacional", "addfoto", "addruta", "addconvenio_dz", "addioarr", "init", "index", "reporte_documental_anual"}:
            if not os.path.isfile(MAESTRA_DEFAULT):
                return False, f"No se encontró el Excel maestro en:\n{MAESTRA_DEFAULT}"

        if action in {"add", "addmantenimiento_grupal", "addchecklist", "addestado_situacional", "addruta", "addconvenio_dz"}:
            src = os.path.realpath(self.src_field.entry.get().strip())
            if not src:
                return False, "Selecciona el archivo fuente"
            if not os.path.isfile(src):
                return False, f"No existe el archivo fuente:\n{src}"

        if action == "addmatricula":
            docs = self.get_matricula_docs()
            if not docs:
                return False, "Agrega al menos un documento de matrícula"
            for path, tipo, nombre in docs:
                if not os.path.isfile(path):
                    return False, f"No existe el documento de matrícula:\n{path}"
                if not nombre.strip():
                    return False, "Cada documento de matrícula debe tener un nombre final"

        if action == "addioarr":
            docs = self.get_ioarr_docs()
            if not docs:
                return False, "Agrega al menos un documento IOARR"
            for path, tipo in docs:
                if not os.path.isfile(path):
                    return False, f"No existe el documento IOARR:\n{path}"
                if not tipo.strip():
                    return False, "Cada documento IOARR debe tener un tipo documental"

        if action == "addfoto":
            if not self.get_photo_paths():
                return False, "Debes cargar al menos un archivo en las casillas."

        if action in {"add", "addmatricula", "addmantenimiento_grupal", "addchecklist", "addestado_situacional", "addfoto", "addruta", "addconvenio_dz", "addioarr"}:
            if not self.fecha_entry.get().strip():
                return False, "Selecciona una fecha"

        if action in {"add", "addmatricula", "addchecklist", "addestado_situacional", "addfoto", "index"}:
            if not self.codigo_var.get().strip():
                return False, "Ingresa el código de estación"

        if action in {"addruta", "addchecklist", "addestado_situacional", "addfoto"} and not self.ruta_var.get().strip():
            return False, "Ingresa la ruta, por ejemplo RUTA_01"

        if action == "index" and not self.index_year_var.get().strip():
            return False, "Ingresa el año para reconstruir el índice"

        if action == "reporte_documental_anual" and not self.index_year_var.get().strip():
            return False, "Ingresa el año para generar el reporte documental anual"

        if action == "add" and self.categoria_var.get() == "CERTIFICADO_COMPROBACION_SENSOR" and not self.variable_sensor_var.get().strip():
            return False, "Selecciona la variable/sensor para el certificado de comprobación"


        if action == "addioarr" and not self.nombre_ioarr_var.get().strip():
            return False, "Ingresa el nombre corto de la IOARR, por ejemplo PUENTE_BRENA-TULUMAYO"

        return True, ""

    def build_command(self):
        action = self.current_action
        src = os.path.realpath(self.src_field.entry.get().strip())

        if action == "addfoto":
            src = self.build_photos_pdf()

        command = ["python", SCRIPT_DEFAULT, action]

        if action == "init":
            command += ["--dz", self.dz_var.get()]
            years = self.years_var.get().strip().split()
            if years:
                command += ["--years", *years]
            selected_codes = self.get_selected_station_codes()
            if selected_codes:
                command += ["--estaciones", ",".join(selected_codes)]
            command += ["--maestra", MAESTRA_DEFAULT]
            if self.overwrite_var.get():
                command += ["--ficha-overwrite"]

        elif action == "add":
            command += [
                "--src", src,
                "--categoria", self.categoria_var.get(),
                "--dz", self.dz_var.get(),
                "--codigo", self.codigo_var.get().strip(),
                "--fecha", self.fecha_entry.get().strip(),
                "--maestra", MAESTRA_DEFAULT
            ]
            if self.categoria_var.get() == "SINIESTROS":
                command += ["--subtipo-siniestro", self.siniestro_subtipo_var.get()]
            if self.categoria_var.get() == "CERTIFICADO_COMPROBACION_SENSOR":
                command += ["--variable-sensor", self.variable_sensor_var.get()]
            if self.copy_mode_var.get() == "copy":
                command += ["--copy"]

        elif action == "addmatricula":
            self.refresh_matricula_final_names()
            docs = self.get_matricula_docs()
            srcs = [d[0] for d in docs]
            tipos = [d[1] for d in docs]
            nombres = [d[2] for d in docs]
            command += ["--srcs", *srcs]
            command += ["--tipos", *tipos]
            command += ["--nombres-finales", *nombres]
            command += [
                "--dz", self.dz_var.get(),
                "--codigo", self.codigo_var.get().strip(),
                "--fecha", self.fecha_entry.get().strip(),
                "--maestra", MAESTRA_DEFAULT
            ]
            if self.copy_mode_var.get() == "copy":
                command += ["--copy"]

        elif action == "addchecklist":
            command += [
                "--src", src,
                "--dz", self.dz_var.get(),
                "--ruta", self.ruta_var.get().strip(),
                "--codigo", self.codigo_var.get().strip(),
                "--fecha", self.fecha_entry.get().strip(),
                "--maestra", MAESTRA_DEFAULT
            ]
            if self.copy_mode_var.get() == "copy":
                command += ["--copy"]

        elif action == "addestado_situacional":
            command += [
                "--src", src,
                "--dz", self.dz_var.get(),
                "--ruta", self.ruta_var.get().strip(),
                "--codigo", self.codigo_var.get().strip(),
                "--fecha", self.fecha_entry.get().strip(),
                "--maestra", MAESTRA_DEFAULT
            ]
            if self.copy_mode_var.get() == "copy":
                command += ["--copy"]

        elif action == "addfoto":
            command += [
                "--src", src,
                "--dz", self.dz_var.get(),
                "--ruta", self.ruta_var.get().strip(),
                "--codigo", self.codigo_var.get().strip(),
                "--fecha", self.fecha_entry.get().strip(),
                "--maestra", MAESTRA_DEFAULT
            ]
            if self.copy_mode_var.get() == "copy":
                command += ["--copy"]

        elif action == "addmantenimiento_grupal":
            command += [
                "--src", src,
                "--dz", self.dz_var.get(),
                "--fecha", self.fecha_entry.get().strip()
            ]
            selected_codes = self.get_selected_station_codes()
            if selected_codes:
                command += ["--estaciones", ",".join(selected_codes)]
            command += ["--maestra", MAESTRA_DEFAULT]
            if self.responsable_var.get().strip():
                command += ["--responsable", self.responsable_var.get().strip()]
            if self.obs_var.get().strip():
                command += ["--obs", self.obs_var.get().strip()]
            if self.copy_mode_var.get() == "copy":
                command += ["--copy"]

        elif action == "addruta":
            command += [
                "--src", src,
                "--dz", self.dz_var.get(),
                "--ruta", self.ruta_var.get().strip(),
                "--tipo", self.tipo_var.get(),
                "--fecha", self.fecha_entry.get().strip()
            ]
            selected_codes = self.get_selected_station_codes()
            if selected_codes:
                command += ["--estaciones", ",".join(selected_codes)]
            command += ["--maestra", MAESTRA_DEFAULT]
            if self.responsable_var.get().strip():
                command += ["--responsable", self.responsable_var.get().strip()]
            if self.obs_var.get().strip():
                command += ["--obs", self.obs_var.get().strip()]
            if self.copy_mode_var.get() == "copy":
                command += ["--copy"]

        elif action == "addconvenio_dz":
            command += [
                "--src", src,
                "--dz", self.dz_var.get(),
                "--fecha", self.fecha_entry.get().strip(),
                "--tipo-documento-convenio", self.tipo_documento_convenio_var.get()
            ]
            selected_codes = self.get_selected_station_codes()
            if selected_codes:
                command += ["--estaciones", ",".join(selected_codes)]
            command += ["--maestra", MAESTRA_DEFAULT]
            if self.obs_var.get().strip():
                command += ["--obs", self.obs_var.get().strip()]
            if self.copy_mode_var.get() == "copy":
                command += ["--copy"]

        elif action == "addioarr":
            docs = self.get_ioarr_docs()
            srcs = [d[0] for d in docs]
            tipos = [d[1] for d in docs]

            command += ["--srcs", *srcs]
            command += ["--tipos-documento", *tipos]
            command += [
                "--dz", self.dz_var.get(),
                "--fecha", self.fecha_entry.get().strip(),
                "--nombre-ioarr", self.nombre_ioarr_var.get().strip(),
            ]
            selected_codes = self.get_selected_station_codes()
            if selected_codes:
                command += ["--estaciones", ",".join(selected_codes)]
            command += ["--maestra", MAESTRA_DEFAULT]
            if self.obs_var.get().strip():
                command += ["--obs", self.obs_var.get().strip()]
            if self.copy_mode_var.get() == "copy":
                command += ["--copy"]

        elif action == "reporte_documental_anual":
            command += [
                "--dz", self.dz_var.get(),
                "--year", self.index_year_var.get().strip(),
                "--maestra", MAESTRA_DEFAULT
            ]

        elif action == "index":
            command += [
                "--dz", self.dz_var.get(),
                "--year", self.index_year_var.get().strip(),
                "--codigo", self.codigo_var.get().strip(),
                "--maestra", MAESTRA_DEFAULT
            ]

        elif action == "addficha_dz":
            command += ["--dz", self.dz_var.get()]
            if self.filename_var.get().strip():
                command += ["--filename", self.filename_var.get().strip()]
            if self.overwrite_var.get():
                command += ["--overwrite"]

        elif action == "mk_ficha_dz":
            command += ["--dz", self.dz_var.get()]

        return command

    def run_command(self):
        if self.is_running:
            messagebox.showinfo("Proceso en ejecución", "Ya hay un proceso ejecutándose.")
            return

        ok, msg = self.validate()
        if not ok:
            messagebox.showwarning("Faltan datos", msg)
            return

        try:
            cmd = self.build_command()
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return

        self.log("\n" + "=" * 100 + "\n")
        self.log(f"Acción: {self.current_action}\n")
        self.log("Ejecutando:\n")
        self.log(" ".join(f'"{x}"' if " " in x else x for x in cmd) + "\n\n")

        decision_copy = self.copy_mode_var.get() == "copy"
        texto_decision = "sí (copiar)" if decision_copy else "no (mover)"
        self.log(f"Respuesta automática a copiar/mover: {texto_decision}\n")

        self.set_running_state(True)
        t = threading.Thread(
            target=self._run_command_background,
            args=(cmd, decision_copy),
            daemon=True
        )
        t.start()

    def _run_command_background(self, cmd, wants_copy):
        try:
            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                stdin=subprocess.PIPE,
                text=True
            )

            respuesta = "s\n" if wants_copy else "n\n"
            texto_respuesta = "s" if wants_copy else "n"

            self.after(0, self.log, f"↳ Respuesta enviada automáticamente: {texto_respuesta}\n")

            try:
                stdout_data, _ = process.communicate(input=respuesta)
            except TypeError:
                process.stdin.write(respuesta)
                process.stdin.flush()
                stdout_data, _ = process.communicate()

            if stdout_data:
                self.after(0, self.log, stdout_data)

            if process.returncode == 0:
                self.after(0, self.log, "\n✅ Proceso finalizado correctamente.\n")
                self.clear_after_success()
                self.after(0, lambda: messagebox.showinfo("Éxito", "Proceso finalizado correctamente."))
            else:
                self.after(0, self.log, f"\n❌ El proceso terminó con código {process.returncode}\n")
                self.after(0, lambda: messagebox.showerror("Error", f"El proceso terminó con código {process.returncode}"))

        except Exception as e:
            self.after(0, self.log, f"\n❌ Error al ejecutar: {e}\n")
            self.after(0, lambda: messagebox.showerror("Error", str(e)))
        finally:
            self.after(0, self.set_running_state, False)


if __name__ == "__main__":
    app = LegajosGUIV2()
    app.mainloop()
