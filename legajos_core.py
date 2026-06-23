#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
LEGAJOS_codigo.py — Gestor de legajos para SGR/DRD SENAMHI

- Fechas de entrada: SOLO DD-MM-YYYY
- Las carpetas de estación se construyen como: CODIGO-CLASIFICACION-NOMBRE
  usando un Excel maestro con columnas: Código, Nombre, Clasificación.
- Incluye MANTENIMIENTO con la estructura:
    - MANTENIMIENTO_INDIVIDUAL (informes, checklist y estado situacional de una estación)
    - RUTA_XX (archivos de checklist, estado situacional y fotos juntos)
    - MANTENIMIENTO_GRUPAL_SIN_RUTA a nivel DZ/año para informes que incluyen varias estaciones sin ruta
- Incluye INSTALACION_ESTACION como sección por estación para actas, fichas e informes de matrícula
- Comandos: init, add, addmatricula, addmantenimiento_grupal, addruta, addconvenio_dz, addchecklist,
            addestado_situacional, addfoto, addficha_dz, reporte_documental_anual, index, mk_ficha_dz.
"""

import argparse
import csv
from datetime import datetime
from pathlib import Path
import shutil
import os
import sys
import re
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
import json

# =========================
# CONFIG
# =========================
#ROOT = Path("LEGAJOS_ESTACION")
#ROOT = Path("/home/sgr05/Drive/DRD_LEGAJOS_ESTACION")
#STATIONS_XLSX_DEFAULT = Path("estaciones_2026-04-21.xlsx")

# =========================
# CONFIG
# =========================

BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"


def load_config():
    default_config = {
        "ROOT": "LEGAJOS_ESTACION",
        "MAESTRA": "MaestraEstaciones.xlsx"
    }

    if not CONFIG_PATH.exists():
        with CONFIG_PATH.open("w", encoding="utf-8") as f:
            json.dump(default_config, f, indent=4, ensure_ascii=False)
        return default_config

    with CONFIG_PATH.open("r", encoding="utf-8") as f:
        config = json.load(f)

    for key, value in default_config.items():
        config.setdefault(key, value)

    return config


CONFIG = load_config()

ROOT = Path(CONFIG["ROOT"]).expanduser()
if not ROOT.is_absolute():
    ROOT = BASE_DIR / ROOT

STATIONS_XLSX_DEFAULT = Path(CONFIG["MAESTRA"]).expanduser()
if not STATIONS_XLSX_DEFAULT.is_absolute():
    STATIONS_XLSX_DEFAULT = BASE_DIR / STATIONS_XLSX_DEFAULT



CATEGORIAS_ESTACION = [
    "MANTENIMIENTO",
    "CHECKLIST_MANTENIMIENTO",
    "ESTADO_SITUACIONAL",
    "INSPECCION",
    "CALIBRACION",
    "AFOROS",
    "CALIDAD_DATOS",
    "INCIDENCIAS",
    "INSTALACIONES_NUEVAS",
    "INSTALACION_ESTACION",
    "INSTALACION_EQUIPO",
    "CESE_OBSERVADOR",
    "SINIESTROS",
    "REUBICACION",
    "SOLICITUD_REUBICACION",
    "BAJA_ESTACION",
    "ACTUALIZACION_METADATA",
]

DOCUMENTAL_CATEGORIES = [
    "INSTALACION_ESTACION",
    "CONVENIOS",
    "CESE_OBSERVADOR",
    "SINIESTROS",
    "REUBICACION",
    "SOLICITUD_REUBICACION",
    "BAJA_ESTACION",
    "ACTUALIZACION_METADATA",
]

DOCUMENTAL_DISPLAY = {
    "INSTALACION_ESTACION": "INSTALACIÓN DE ESTACIÓN",
    "INSTALACION_EQUIPO": "INSTALACIÓN DE EQUIPO",
    "CONVENIOS": "CONVENIO",
    "CESE_OBSERVADOR": "CESE DE OBSERVADOR",
    "SINIESTROS": "SINIESTRO",
    "REUBICACION": "REUBICACIÓN",
    "SOLICITUD_REUBICACION": "SOLICITUD DE REUBICACIÓN",
    "BAJA_ESTACION": "BAJA DE ESTACIÓN",
    "ACTUALIZACION_METADATA": "ACTUALIZACIÓN DE METADATA",
}

SINIESTRO_SUBTIPOS = [
    "ROBO",
    "DAÑO_POR_TERCEROS",
    "DAÑO_POR_FENOMENO_NATURAL",
    "ACCIDENTE",
    "OTRO",
]

REPORTE_DOCUMENTAL_ANUAL_DIRNAME = "REPORTE_DOCUMENTAL_ANUAL"

ESTADO_SITUACIONAL_LABEL = "ESTADO SITUACIONAL DE LA ESTACIÓN PREVIA AL MANTENIMIENTO"
MANTENIMIENTO_INDIVIDUAL_DIRNAME = "MANTENIMIENTO_INDIVIDUAL"
MANTENIMIENTO_GRUPAL_DIRNAME = "MANTENIMIENTO_GRUPAL_SIN_RUTA"

RUTAS_TIPOS = ["MANTENIMIENTOS", "AFOROS", "INSPECCION"]
RUTA_SINGULAR = {"MANTENIMIENTOS": "MANTENIMIENTO", "AFOROS": "AFORO", "INSPECCION": "INSPECCION"}

DZ_CONVENIOS_DIRNAME = "CONVENIOS"
FICHA_MATRICULA_DIRNAME = "Ficha de Matricula"

MATRICULA_DOC_TYPES = [
    "ACTA_INSTALACION",
    "FICHA_MATRICULA",
    "INFORME_MATRICULA",
    "INFORME_INSTALACION",
    "CONVENIO",
    "ANEXO",
    "OTRO",
]

FICHA_COLUMNS = [
    "CODIGO", "NOMBRE_ESTACION", "DZ", "DEPARTAMENTO", "PROVINCIA", "DISTRITO",
    "LATITUD", "LONGITUD", "ALTITUD_m",
    "TIPO_ESTACION", "CLASIFICACION", "CATEGORIA_APLICACION",
    "FECHA_INSTALACION", "ESTADO", "RESPONSABLE",
    "ENTIDAD", "EQUIPAMIENTO", "SERIE", "MARCA", "MODELO",
    "WIGOS_ID", "CODIGO_OMM", "CODIGO_INTERNO",
    "OBSERVACIONES"
]

DATE_INPUT_FMT = "%d-%m-%Y"
NS_MAIN = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

# =========================
# UTILIDADES
# =========================
def slug(s: str) -> str:
    s = s.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    s = s.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    s = s.strip().replace(" ", "_")
    s = re.sub(r'["\'\(\)\[\]\{\}:;,]', '', s)
    return s


def normalize_casefold(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii").casefold()


def ensure_dirs(p: Path):
    p.mkdir(parents=True, exist_ok=True)


def parse_fecha_ddmmyyyy(fecha_str: str) -> datetime:
    try:
        return datetime.strptime(fecha_str, DATE_INPUT_FMT)
    except ValueError:
        raise ValueError("⚠️ Fecha inválida. Usa el formato DD-MM-YYYY (ejemplo: 05-03-2025).")


def prompt_fecha_ddmmyyyy(label: str = "Fecha (DD-MM-YYYY)") -> str:
    while True:
        raw = input(f"{label}: ").strip()
        try:
            _ = parse_fecha_ddmmyyyy(raw)
            return raw
        except ValueError as e:
            print(str(e))

# =========================
# EXCEL MAESTRO
# =========================
def _xlsx_col_letters_to_index(col_letters: str) -> int:
    idx = 0
    for ch in col_letters:
        idx = idx * 26 + (ord(ch.upper()) - ord('A') + 1)
    return idx - 1


def _xlsx_parse_cell_ref(cell_ref: str):
    m = re.match(r"([A-Z]+)(\d+)$", cell_ref)
    if not m:
        raise ValueError(f"Referencia de celda inválida: {cell_ref}")
    return _xlsx_col_letters_to_index(m.group(1)), int(m.group(2))


def _xlsx_read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    out = []
    for si in root.findall("a:si", NS_MAIN):
        texts = [t.text or "" for t in si.findall(".//a:t", NS_MAIN)]
        out.append("".join(texts))
    return out


def _xlsx_read_sheet_rows(xlsx_path: Path, sheet_name: str | None = None) -> list[list[str]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"No existe el Excel maestro: {xlsx_path}")

    with zipfile.ZipFile(xlsx_path) as zf:
        shared_strings = _xlsx_read_shared_strings(zf)

        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rel_root if rel.tag.endswith("Relationship")}

        target_sheet = None
        for s in wb_root.find("a:sheets", NS_MAIN):
            s_name = s.attrib.get("name")
            r_id = s.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if sheet_name is None or s_name == sheet_name:
                target_sheet = rel_map[r_id]
                break

        if target_sheet is None:
            raise ValueError(f"No se encontró una hoja válida en {xlsx_path}")

        sheet_xml_path = "xl/" + target_sheet.lstrip("/")
        sheet_root = ET.fromstring(zf.read(sheet_xml_path))
        sheet_data = sheet_root.find("a:sheetData", NS_MAIN)

        rows = []
        for row in sheet_data:
            row_map = {}
            max_col = -1
            for c in row.findall("a:c", NS_MAIN):
                ref = c.attrib.get("r")
                col_idx, _ = _xlsx_parse_cell_ref(ref)
                max_col = max(max_col, col_idx)
                t = c.attrib.get("t")
                v = c.find("a:v", NS_MAIN)
                inline = c.find("a:is", NS_MAIN)
                value = ""
                if t == "s" and v is not None:
                    value = shared_strings[int(v.text)]
                elif t == "inlineStr" and inline is not None:
                    value = "".join(tn.text or "" for tn in inline.findall(".//a:t", NS_MAIN))
                elif v is not None:
                    value = v.text or ""
                row_map[col_idx] = value
            if max_col >= 0:
                rows.append([row_map.get(i, "") for i in range(max_col + 1)])
    return rows


def build_station_folder_name(codigo: str, clasificacion: str, nombre: str) -> str:
    return slug("-".join([p.strip() for p in [codigo, clasificacion, nombre] if p and p.strip()]))


def load_station_catalog(xlsx_path: Path) -> dict[str, dict[str, str]]:
    rows = _xlsx_read_sheet_rows(xlsx_path)
    if not rows:
        raise ValueError(f"El Excel maestro está vacío: {xlsx_path}")

    headers = rows[0]
    header_map = {normalize_casefold(h): i for i, h in enumerate(headers)}
    required = ["codigo", "nombre", "clasificacion"]
    missing = [h for h in required if h not in header_map]
    if missing:
        raise ValueError(f"El Excel maestro debe contener las columnas Código, Nombre y Clasificación. Faltan: {missing}")

    idx_codigo = header_map["codigo"]
    idx_nombre = header_map["nombre"]
    idx_clas = header_map["clasificacion"]

    catalog = {}
    for row in rows[1:]:
        codigo = (row[idx_codigo] if idx_codigo < len(row) else "").strip()
        nombre = (row[idx_nombre] if idx_nombre < len(row) else "").strip()
        clasificacion = (row[idx_clas] if idx_clas < len(row) else "").strip()
        if not codigo:
            continue
        catalog[codigo] = {
            "codigo": codigo,
            "nombre": nombre,
            "clasificacion": clasificacion,
            "folder_name": build_station_folder_name(codigo, clasificacion, nombre),
        }
    return catalog


def get_station_meta(codigo: str, xlsx_path: Path) -> dict[str, str]:
    catalog = load_station_catalog(xlsx_path)
    codigo = codigo.strip()
    if codigo not in catalog:
        raise ValueError(f"El código de estación '{codigo}' no existe en el Excel maestro: {xlsx_path}")
    return catalog[codigo]

# =========================
# RUTAS Y CARPETAS
# =========================
def dz_dir(dz: str) -> Path:
    return ROOT / dz.upper()


def year_dir(dz: str, year: int | str) -> Path:
    return dz_dir(dz) / f"{int(year):04d}"


def estacion_dir(dz: str, year: int | str, station_meta_or_folder) -> Path:
    folder = station_meta_or_folder["folder_name"] if isinstance(station_meta_or_folder, dict) else str(station_meta_or_folder)
    return year_dir(dz, year) / folder


def cat_dir_estacion(dz: str, year: int | str, station_meta_or_folder, categoria: str) -> Path:
    return estacion_dir(dz, year, station_meta_or_folder) / categoria


def mantenimiento_correctivo_dir_estacion(dz: str, year: int | str, station_meta_or_folder) -> Path:
    return cat_dir_estacion(dz, year, station_meta_or_folder, "MANTENIMIENTO") / MANTENIMIENTO_INDIVIDUAL_DIRNAME


def mantenimiento_ruta_dir_estacion(dz: str, year: int | str, station_meta_or_folder, ruta: str) -> Path:
    return cat_dir_estacion(dz, year, station_meta_or_folder, "MANTENIMIENTO") / ruta.upper().strip()


def rutas_dir(dz: str, year: int | str) -> Path:
    return year_dir(dz, year) / "RUTAS"


def mantenimiento_grupal_dir(dz: str, year: int | str) -> Path:
    return year_dir(dz, year) / MANTENIMIENTO_GRUPAL_DIRNAME


def rutas_cat_dir(dz: str, year: int | str, tipo: str) -> Path:
    return rutas_dir(dz, year) / tipo


def dz_convenios_dir(dz: str) -> Path:
    return dz_dir(dz) / DZ_CONVENIOS_DIRNAME


def dz_ficha_matricula_dir(dz: str) -> Path:
    return dz_dir(dz) / FICHA_MATRICULA_DIRNAME


def reporte_documental_anual_dir(dz: str, year: int | str) -> Path:
    return year_dir(dz, year) / REPORTE_DOCUMENTAL_ANUAL_DIRNAME


def reporte_documental_anual_path(dz: str, year: int | str) -> Path:
    return reporte_documental_anual_dir(dz, year) / f"reporte_documental_anual_{dz.upper()}_{int(year):04d}.xlsx"

# =========================
# ÍNDICES / XLSX
# =========================
def _apply_xlsx_style(ws):
    """Aplica formato básico para que los índices se vean mejor en Google Drive/Sheets."""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 55)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass


def write_xlsx_if_missing(path: Path, headers):
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "INDICE"
    ws.append(headers)
    _apply_xlsx_style(ws)
    wb.save(path)


def append_xlsx_row(path: Path, row: list):
    from openpyxl import load_workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo XLSX: {path}")

    wb = load_workbook(path)
    ws = wb.active
    ws.append(row)
    _apply_xlsx_style(ws)
    wb.save(path)


def rewrite_xlsx(path: Path, headers: list, rows: list[list]):
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "INDICE"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _apply_xlsx_style(ws)
    wb.save(path)


def read_xlsx_as_dicts(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    if not path.exists():
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        out.append({headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))})
    return out


def read_legacy_csv_as_dicts(path: Path) -> list[dict]:
    """Solo para recuperar información si todavía queda algún CSV antiguo."""
    if not path.exists():
        return []
    try:
        with path.open("r", newline="", encoding="utf-8") as fh:
            return list(csv.DictReader(fh))
    except Exception:
        return []


def mantenimiento_grupal_index_path(dz: str, year: int | str) -> Path:
    return mantenimiento_grupal_dir(dz, year) / f"legajo_mantenimiento_grupal_index_{dz.upper()}_{int(year):04d}.xlsx"


def mantenimiento_grupal_index_path_legacy(dz: str, year: int | str) -> Path:
    return mantenimiento_grupal_dir(dz, year) / "legajo_mantenimiento_grupal_index.xlsx"


def ensure_mantenimiento_grupal_indices(dz: str, year: int | str):
    headers = ["DZ", "Año", "Archivo", "Path", "Estaciones_incluidas", "Fecha_ejecucion", "Responsable", "Observaciones"]
    write_xlsx_if_missing(mantenimiento_grupal_index_path(dz, year), headers)
    write_xlsx_if_missing(mantenimiento_grupal_index_path_legacy(dz, year), headers)


def append_mantenimiento_grupal_index_rows(dz: str, year: int | str, row: list):
    ensure_mantenimiento_grupal_indices(dz, year)
    append_xlsx_row(mantenimiento_grupal_index_path(dz, year), row)
    append_xlsx_row(mantenimiento_grupal_index_path_legacy(dz, year), row)


def rutas_index_path(dz: str, year: int | str) -> Path:
    return rutas_dir(dz, year) / f"legajo_rutas_index_{dz.upper()}_{int(year):04d}.xlsx"


def rutas_index_path_legacy(dz: str, year: int | str) -> Path:
    return rutas_dir(dz, year) / "legajo_rutas_index.xlsx"


def ensure_rutas_indices(dz: str, year: int | str):
    headers = ["DZ", "Ruta", "Año", "Tipo", "Archivo", "Path", "Estaciones_incluidas", "Fecha_ejecucion", "Responsable", "Observaciones"]
    write_xlsx_if_missing(rutas_index_path(dz, year), headers)
    write_xlsx_if_missing(rutas_index_path_legacy(dz, year), headers)


def append_rutas_index_rows(dz: str, year: int | str, row: list):
    ensure_rutas_indices(dz, year)
    append_xlsx_row(rutas_index_path(dz, year), row)
    append_xlsx_row(rutas_index_path_legacy(dz, year), row)


def estaciones_rutas_path(dz: str) -> Path:
    return dz_dir(dz) / "estaciones_rutas.xlsx"


def dz_convenios_index_path(dz: str) -> Path:
    return dz_convenios_dir(dz) / f"legajo_convenios_dz_{dz.upper()}.xlsx"


def dz_convenios_index_path_legacy(dz: str) -> Path:
    return dz_convenios_dir(dz) / "legajo_convenios_dz.xlsx"


def ensure_dz_convenios_indices(dz: str):
    headers = ["dz", "filename", "stored_at", "fecha", "estaciones_incluidas", "observaciones"]
    write_xlsx_if_missing(dz_convenios_index_path(dz), headers)
    write_xlsx_if_missing(dz_convenios_index_path_legacy(dz), headers)


def append_dz_convenios_row(dz: str, row: list):
    ensure_dz_convenios_indices(dz)
    append_xlsx_row(dz_convenios_index_path(dz), row)
    append_xlsx_row(dz_convenios_index_path_legacy(dz), row)

# =========================
# NOMBRES DE ARCHIVO
# =========================
def build_filename_estacion(categoria: str, dz: str, station_meta: dict[str, str], fecha: datetime, ext: str) -> str:
    return f"{categoria}_{dz.upper()}_{station_meta['folder_name']}_{fecha.date()}.{ext.lstrip('.').lower()}"


def build_filename_ruta(ruta: str, tipo: str, fecha: datetime, ext: str) -> str:
    tipo_singular = RUTA_SINGULAR.get(tipo, tipo)
    return f"{ruta}_{tipo_singular}_{fecha.date()}.{ext.lstrip('.').lower()}"


def build_filename_mantenimiento_grupal(dz: str, src_stem: str, fecha: datetime, ext: str) -> str:
    nombre_base = slug(src_stem) or "INFORME"
    return f"MANTENIMIENTO_GRUPAL_SIN_RUTA_{dz.upper()}_{nombre_base}_{fecha.date()}.{ext.lstrip('.').lower()}"


def build_filename_checklist(ruta: str, dz: str, station_meta: dict[str, str], fecha: datetime, ext: str) -> str:
    return f"CHECKLIST_MANTENIMIENTO_{ruta.upper()}_{dz.upper()}_{station_meta['folder_name']}_{fecha.date()}.{ext.lstrip('.').lower()}"


def build_filename_estado_situacional(ruta: str, dz: str, station_meta: dict[str, str], fecha: datetime, ext: str) -> str:
    return f"{slug(ESTADO_SITUACIONAL_LABEL)}_{ruta.upper()}_{dz.upper()}_{station_meta['folder_name']}_{fecha.date()}.{ext.lstrip('.').lower()}"


def build_filename_foto(ruta: str, dz: str, station_meta: dict[str, str], fecha: datetime, ext: str) -> str:
    return f"FOTO_MANTENIMIENTO_{ruta.upper()}_{dz.upper()}_{station_meta['folder_name']}_{fecha.date()}.{ext.lstrip('.').lower()}"


def build_filename_ficha(dz: str) -> str:
    return f"Ficha_de_Matricula_{dz.upper()}.xlsx"


def clean_output_filename(name: str, default_ext: str = "pdf") -> str:
    """Limpia un nombre final ingresado desde la app, conservando la extensión."""
    name = (name or "").strip()
    if not name:
        name = f"DOCUMENTO.{default_ext.lstrip('.')}"

    stem = Path(name).stem
    ext = Path(name).suffix.lower().lstrip('.') or default_ext.lstrip('.').lower()
    stem = slug(stem) or "DOCUMENTO"
    return f"{stem}.{ext}"


def build_filename_matricula_documento(tipo_doc: str, dz: str, station_meta: dict[str, str], fecha: datetime, ext: str, orden: int | None = None) -> str:
    tipo = (tipo_doc or "OTRO").upper().strip()
    if tipo not in MATRICULA_DOC_TYPES:
        tipo = "OTRO"
    suffix = f"_{orden:02d}" if orden is not None else ""
    return f"{tipo}_{dz.upper()}_{station_meta['folder_name']}_{fecha.date()}{suffix}.{ext.lstrip('.').lower()}"

# =========================
# README / INDEX
# =========================
def write_estacion_readme(dz: str, year: int | str, station_meta: dict[str, str]):
    d = estacion_dir(dz, year, station_meta)
    ensure_dirs(d)
    txt = (
        "# Legajo de estación (por año)\n"
        f"- DZ: {dz.upper()}\n"
        f"- Año: {int(year):04d}\n"
        f"- Código: {station_meta['codigo']}\n"
        f"- Clasificación: {station_meta['clasificacion']}\n"
        f"- Nombre: {station_meta['nombre']}\n"
        f"- Carpeta estación: {station_meta['folder_name']}\n"
        "- Estructura: AAAA/ESTACION/CATEGORIA/archivo\n"
        f"- MANTENIMIENTO usa: '{MANTENIMIENTO_INDIVIDUAL_DIRNAME}' para mantenimientos individuales, checklist y estado situacional por estación y 'RUTA_XX' para evidencias de ruta. Los informes grupales sin ruta se guardan en DZ/AÑO/{MANTENIMIENTO_GRUPAL_DIRNAME}.\n"
        "- Importante: al ingresar fechas use siempre DD-MM-YYYY.\n"
    )
    (d / "README.md").write_text(txt, encoding="utf-8")


def scan_files_estacion(dz: str, year: int | str, station_meta: dict[str, str]):
    base = estacion_dir(dz, year, station_meta)
    if not base.exists():
        return []
    rows = []
    for cat in CATEGORIAS_ESTACION:
        cdir = base / cat
        if not cdir.exists():
            continue
        if cat in {"MANTENIMIENTO", "SINIESTROS"}:
            for f in sorted([p for p in cdir.rglob("*") if p.is_file()]):
                rows.append((cat, f))
        else:
            for f in sorted([p for p in cdir.iterdir() if p.is_file()]):
                rows.append((cat, f))
    return rows


def update_legajo_index_estacion(dz: str, year: int | str, station_meta: dict[str, str]):
    base = estacion_dir(dz, year, station_meta)
    ensure_dirs(base)

    headers = ["categoria", "year", "dz", "estacion_slug", "filename", "relpath"]
    prev_external = []

    # Recupera referencias externas desde el XLSX actual y, si aún existe, desde el CSV antiguo.
    possible_previous_files = [
        base / "legajo_index.xlsx",
        base / "legajo_index.csv",
    ]

    for previous_path in possible_previous_files:
        if previous_path.suffix.lower() == ".xlsx":
            previous_rows = read_xlsx_as_dicts(previous_path)
        else:
            previous_rows = read_legacy_csv_as_dicts(previous_path)

        for row in previous_rows:
            rel = str(row.get("relpath", "") or "")
            if rel.startswith("..") or "/RUTAS/" in rel or "/CONVENIOS/" in rel:
                prev_external.append([
                    row.get("categoria", ""),
                    row.get("year", ""),
                    row.get("dz", ""),
                    row.get("estacion_slug", ""),
                    row.get("filename", ""),
                    rel,
                ])

    rows_local = []
    for cat, f in scan_files_estacion(dz, year, station_meta):
        # Evita que el índice se registre a sí mismo como documento de la estación.
        if f.suffix.lower() == ".xlsx" and f.name.startswith("legajo_index"):
            continue
        rel = os.path.relpath(f, start=base)
        rows_local.append([cat, f"{int(year):04d}", dz.upper(), station_meta["folder_name"], f.name, rel])

    merged = rows_local + prev_external
    seen = set()
    deduped = []
    for cat, yr, dzv, est_slug, fname, rel in merged:
        key = (cat, fname, rel)
        if key not in seen:
            seen.add(key)
            deduped.append([cat, yr, dzv, est_slug, fname, rel])

    legacy_path = base / "legajo_index.xlsx"
    explicit_path = base / f"legajo_index_{station_meta['folder_name']}_{int(year):04d}.xlsx"

    rewrite_xlsx(legacy_path, headers, deduped)
    rewrite_xlsx(explicit_path, headers, deduped)


def add_reference_to_stations(dz: str, year: int | str, estaciones_meta: list[dict[str, str]], categoria_ref: str, archivo: str, rel_from_station: str):
    headers = ["categoria", "year", "dz", "estacion_slug", "filename", "relpath"]

    for meta in estaciones_meta:
        base = estacion_dir(dz, year, meta)
        ensure_dirs(base)
        xlsx_path = base / "legajo_index.xlsx"
        write_xlsx_if_missing(xlsx_path, headers)
        append_xlsx_row(xlsx_path, [categoria_ref, f"{int(year):04d}", dz.upper(), meta["folder_name"], archivo, rel_from_station])

        explicit_path = base / f"legajo_index_{meta['folder_name']}_{int(year):04d}.xlsx"
        write_xlsx_if_missing(explicit_path, headers)
        append_xlsx_row(explicit_path, [categoria_ref, f"{int(year):04d}", dz.upper(), meta["folder_name"], archivo, rel_from_station])

# =========================
# OPERACIONES
# =========================
def add_report_estacion(src: Path, categoria: str, dz: str, codigo: str, fecha_str_ddmmyyyy: str, maestra_xlsx: Path, copy=False, subtipo_siniestro: str = ""):
    categoria = categoria.upper().strip()
    if categoria not in CATEGORIAS_ESTACION:
        raise ValueError(f"Categoria no válida: {categoria}. Use: {', '.join(CATEGORIAS_ESTACION)}")
    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo fuente: {src}")

    station_meta = get_station_meta(codigo, maestra_xlsx)
    fecha = parse_fecha_ddmmyyyy(fecha_str_ddmmyyyy)
    year = fecha.year
    ext = src.suffix[1:] if src.suffix else "pdf"

    if categoria in {"MANTENIMIENTO", "CHECKLIST_MANTENIMIENTO", "ESTADO_SITUACIONAL"}:
        dest_dir = mantenimiento_correctivo_dir_estacion(dz, year, station_meta)
    elif categoria == "SINIESTROS":
        subtipo = (subtipo_siniestro or "OTRO").upper().strip()
        if subtipo not in SINIESTRO_SUBTIPOS:
            subtipo = "OTRO"
        dest_dir = cat_dir_estacion(dz, year, station_meta, categoria) / subtipo
    else:
        dest_dir = cat_dir_estacion(dz, year, station_meta, categoria)
    ensure_dirs(dest_dir)
    dst = dest_dir / build_filename_estacion(categoria, dz, station_meta, fecha, ext)

    if copy:
        shutil.copy2(src, dst)
    else:
        shutil.move(str(src), str(dst))

    write_estacion_readme(dz, year, station_meta)
    update_legajo_index_estacion(dz, year, station_meta)
    return dst


def add_documentos_matricula(srcs: list[Path], tipos: list[str], nombres_finales: list[str] | None,
                            dz: str, codigo: str, fecha_str_ddmmyyyy: str,
                            maestra_xlsx: Path, copy=False) -> list[Path]:
    """Guarda uno o varios documentos dentro de ESTACION/INSTALACION_ESTACION/.

    Cada documento puede tener un tipo técnico y un nombre final editable desde la app.
    """
    if not srcs:
        raise ValueError("Debes indicar al menos un documento de matrícula.")

    if len(tipos) != len(srcs):
        raise ValueError("La cantidad de tipos debe coincidir con la cantidad de archivos.")

    if nombres_finales is not None and len(nombres_finales) != len(srcs):
        raise ValueError("La cantidad de nombres finales debe coincidir con la cantidad de archivos.")

    station_meta = get_station_meta(codigo, maestra_xlsx)
    fecha = parse_fecha_ddmmyyyy(fecha_str_ddmmyyyy)
    year = fecha.year
    dest_dir = cat_dir_estacion(dz, year, station_meta, "INSTALACION_ESTACION")
    ensure_dirs(dest_dir)

    used_names = set()
    destinos = []

    for i, src in enumerate(srcs, start=1):
        src = Path(src).expanduser()
        if not src.exists():
            raise FileNotFoundError(f"No existe el archivo fuente: {src}")

        tipo_doc = (tipos[i - 1] or "OTRO").upper().strip()
        if tipo_doc not in MATRICULA_DOC_TYPES:
            tipo_doc = "OTRO"

        ext = src.suffix[1:] if src.suffix else "pdf"

        if nombres_finales and nombres_finales[i - 1].strip():
            fname = clean_output_filename(nombres_finales[i - 1], ext)
        else:
            fname = build_filename_matricula_documento(tipo_doc, dz, station_meta, fecha, ext)

        # Evita sobrescrituras cuando dos filas terminan con el mismo nombre.
        original_stem = Path(fname).stem
        original_ext = Path(fname).suffix
        counter = 2
        while fname in used_names or (dest_dir / fname).exists():
            fname = f"{original_stem}_{counter:02d}{original_ext}"
            counter += 1

        used_names.add(fname)
        dst = dest_dir / fname

        if copy:
            shutil.copy2(src, dst)
        else:
            shutil.move(str(src), str(dst))

        destinos.append(dst)

    write_estacion_readme(dz, year, station_meta)
    update_legajo_index_estacion(dz, year, station_meta)
    return destinos


def add_report_ruta(src: Path, dz: str, ruta: str, tipo: str, fecha_str_ddmmyyyy: str,
                    codigos_estaciones: list[str] | None = None, maestra_xlsx: Path = STATIONS_XLSX_DEFAULT,
                    responsable: str = "", obs: str = "", copy=False):
    if tipo not in RUTAS_TIPOS:
        raise ValueError(f"Tipo de ruta no válido: {tipo}. Usa: {', '.join(RUTAS_TIPOS)}")
    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo fuente: {src}")

    fecha = parse_fecha_ddmmyyyy(fecha_str_ddmmyyyy)
    year = fecha.year
    ext = src.suffix[1:] if src.suffix else "pdf"

    dest_dir = rutas_cat_dir(dz, year, tipo)
    ensure_dirs(dest_dir)

    fname = build_filename_ruta(ruta, tipo, fecha, ext)
    dst = dest_dir / fname
    if copy:
        shutil.copy2(src, dst)
    else:
        shutil.move(str(src), str(dst))

    estaciones_meta = [get_station_meta(c, maestra_xlsx) for c in (codigos_estaciones or [])]
    estaciones_str = ", ".join(meta["folder_name"] for meta in estaciones_meta)
    relpath = os.path.relpath(dst, start=rutas_dir(dz, year))
    row = [dz.upper(), ruta, f"{int(year):04d}", tipo, fname, relpath, estaciones_str, fecha_str_ddmmyyyy, responsable, obs]
    append_rutas_index_rows(dz, year, row)

    if estaciones_meta:
        for meta in estaciones_meta:
            base_est = estacion_dir(dz, year, meta)
            ensure_dirs(base_est)
            write_xlsx_if_missing(base_est / "legajo_index.xlsx", ["categoria", "year", "dz", "estacion_slug", "filename", "relpath"])
            rel_from_station = os.path.relpath(dst, start=base_est)
            add_reference_to_stations(dz, year, [meta], RUTA_SINGULAR.get(tipo, tipo).upper(), fname, rel_from_station)

    return dst



def add_mantenimiento_grupal(src: Path, dz: str, fecha_str_ddmmyyyy: str,
                             codigos_estaciones: list[str] | None = None,
                             maestra_xlsx: Path = STATIONS_XLSX_DEFAULT,
                             responsable: str = "", obs: str = "", copy=False):
    """
    Registra un informe de mantenimiento que involucra varias estaciones,
    sin obligarlo a pertenecer a una ruta.

    Guarda un solo archivo a nivel DZ/año/MANTENIMIENTO_GRUPAL_SIN_RUTA y agrega
    una referencia en el índice de cada estación seleccionada.
    """
    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo fuente: {src}")

    fecha = parse_fecha_ddmmyyyy(fecha_str_ddmmyyyy)
    year = fecha.year
    ext = src.suffix[1:] if src.suffix else "pdf"

    dest_dir = mantenimiento_grupal_dir(dz, year)
    ensure_dirs(dest_dir)
    ensure_mantenimiento_grupal_indices(dz, year)

    fname = build_filename_mantenimiento_grupal(dz, src.stem, fecha, ext)
    dst = dest_dir / fname

    if copy:
        shutil.copy2(src, dst)
    else:
        shutil.move(str(src), str(dst))

    estaciones_meta = [get_station_meta(c, maestra_xlsx) for c in (codigos_estaciones or [])]
    estaciones_str = ", ".join(meta["folder_name"] for meta in estaciones_meta)
    relpath = os.path.relpath(dst, start=mantenimiento_grupal_dir(dz, year))
    row = [dz.upper(), f"{int(year):04d}", fname, relpath, estaciones_str, fecha_str_ddmmyyyy, responsable, obs]
    append_mantenimiento_grupal_index_rows(dz, year, row)

    if estaciones_meta:
        for meta in estaciones_meta:
            base_est = estacion_dir(dz, year, meta)
            ensure_dirs(base_est)
            write_estacion_readme(dz, year, meta)
            rel_from_station = os.path.relpath(dst, start=base_est)
            add_reference_to_stations(dz, year, [meta], "MANTENIMIENTO_GRUPAL_SIN_RUTA", fname, rel_from_station)

    return dst

def add_convenio_dz(src: Path, dz: str, fecha_str_ddmmyyyy: str, codigos_estaciones: list[str] | None = None,
                    maestra_xlsx: Path = STATIONS_XLSX_DEFAULT, obs: str = "", copy: bool = False):
    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo fuente: {src}")

    ensure_dirs(dz_convenios_dir(dz))
    ensure_dz_convenios_indices(dz)

    fecha = parse_fecha_ddmmyyyy(fecha_str_ddmmyyyy)
    year = fecha.year
    ext = src.suffix[1:] if src.suffix else "pdf"
    fname = f"CONVENIO_{dz.upper()}_{fecha.date()}.{ext}"
    dst = dz_convenios_dir(dz) / fname

    if copy:
        shutil.copy2(src, dst)
    else:
        shutil.move(str(src), str(dst))

    estaciones_meta = [get_station_meta(c, maestra_xlsx) for c in (codigos_estaciones or [])]
    estaciones_str = ", ".join(meta["folder_name"] for meta in estaciones_meta)
    rel_from_dz = os.fspath(dst.relative_to(dz_dir(dz)))
    append_dz_convenios_row(dz, [dz.upper(), fname, rel_from_dz, fecha_str_ddmmyyyy, estaciones_str, obs])

    if estaciones_meta:
        for meta in estaciones_meta:
            base_est = estacion_dir(dz, year, meta)
            ensure_dirs(base_est)
            rel_from_station = os.path.relpath(dst, start=base_est)
            add_reference_to_stations(dz, year, [meta], "CONVENIOS", fname, rel_from_station)

    return dst


def add_checklist_estacion(src: Path, dz: str, ruta: str, codigo: str, fecha_str_ddmmyyyy: str, maestra_xlsx: Path, copy=False):
    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo fuente: {src}")
    station_meta = get_station_meta(codigo, maestra_xlsx)
    fecha = parse_fecha_ddmmyyyy(fecha_str_ddmmyyyy)
    year = fecha.year
    ext = src.suffix[1:] if src.suffix else "pdf"

    dest_dir = mantenimiento_ruta_dir_estacion(dz, year, station_meta, ruta)
    ensure_dirs(dest_dir)
    dst = dest_dir / build_filename_checklist(ruta, dz, station_meta, fecha, ext)

    if copy:
        shutil.copy2(src, dst)
    else:
        shutil.move(str(src), str(dst))

    write_estacion_readme(dz, year, station_meta)
    update_legajo_index_estacion(dz, year, station_meta)
    return dst


def add_estado_situacional_estacion(src: Path, dz: str, ruta: str, codigo: str, fecha_str_ddmmyyyy: str, maestra_xlsx: Path, copy=False):
    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo fuente: {src}")
    station_meta = get_station_meta(codigo, maestra_xlsx)
    fecha = parse_fecha_ddmmyyyy(fecha_str_ddmmyyyy)
    year = fecha.year
    ext = src.suffix[1:] if src.suffix else "pdf"

    dest_dir = mantenimiento_ruta_dir_estacion(dz, year, station_meta, ruta)
    ensure_dirs(dest_dir)
    dst = dest_dir / build_filename_estado_situacional(ruta, dz, station_meta, fecha, ext)

    if copy:
        shutil.copy2(src, dst)
    else:
        shutil.move(str(src), str(dst))

    write_estacion_readme(dz, year, station_meta)
    update_legajo_index_estacion(dz, year, station_meta)
    return dst


def add_foto_estacion(src: Path, dz: str, ruta: str, codigo: str, fecha_str_ddmmyyyy: str, maestra_xlsx: Path, copy=False):
    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo fuente: {src}")
    station_meta = get_station_meta(codigo, maestra_xlsx)
    fecha = parse_fecha_ddmmyyyy(fecha_str_ddmmyyyy)
    year = fecha.year
    ext = src.suffix[1:] if src.suffix else "jpg"

    dest_dir = mantenimiento_ruta_dir_estacion(dz, year, station_meta, ruta)
    ensure_dirs(dest_dir)
    dst = dest_dir / build_filename_foto(ruta, dz, station_meta, fecha, ext)

    if copy:
        shutil.copy2(src, dst)
    else:
        shutil.move(str(src), str(dst))

    write_estacion_readme(dz, year, station_meta)
    update_legajo_index_estacion(dz, year, station_meta)
    return dst


def create_ficha_excel(dz: str, filename: str | None = None, overwrite: bool = False) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except Exception as e:
        raise RuntimeError(
            "Se requiere 'openpyxl' para crear el Excel. Instálalo con:\n"
            "  conda install -c conda-forge openpyxl\n"
            "o\n"
            "  pip install openpyxl"
        ) from e

    dzp = dz.upper()
    target_dir = dz_ficha_matricula_dir(dzp)
    ensure_dirs(target_dir)
    fname = filename or build_filename_ficha(dzp)
    outpath = target_dir / fname
    if outpath.exists() and not overwrite:
        return outpath

    wb = Workbook()
    ws = wb.active
    ws.title = "MATRICULA"
    for col_idx, col_name in enumerate(FICHA_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        try:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        except Exception:
            pass
    for col_idx, col_name in enumerate(FICHA_COLUMNS, start=1):
        width = max(12, min(28, len(col_name) + 2))
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"
    wb.save(outpath)
    return outpath


# =========================
# REPORTE DOCUMENTAL ANUAL
# =========================
def _station_meta_from_folder(folder_name: str, catalog_by_folder: dict[str, dict[str, str]]) -> dict[str, str]:
    if folder_name in catalog_by_folder:
        return catalog_by_folder[folder_name]

    parts = folder_name.split("-", 2)
    codigo = parts[0] if len(parts) > 0 else ""
    clasificacion = parts[1] if len(parts) > 1 else ""
    nombre = parts[2].replace("_", " ") if len(parts) > 2 else folder_name
    return {
        "codigo": codigo,
        "nombre": nombre,
        "clasificacion": clasificacion,
        "folder_name": folder_name,
    }


def _is_station_folder(path: Path) -> bool:
    if not path.is_dir():
        return False
    excluded = {
        "RUTAS",
        MANTENIMIENTO_GRUPAL_DIRNAME,
        REPORTE_DOCUMENTAL_ANUAL_DIRNAME,
    }
    if path.name in excluded:
        return False
    return True


def _append_doc_rows(rows_by_category: dict[str, list[list]], resumen_keys: set, year: int, dz: str,
                     meta: dict[str, str], categoria: str, archivo: str, ruta: str, subtipo: str = ""):
    condicion = DOCUMENTAL_DISPLAY.get(categoria, categoria)
    resumen_keys.add((year, dz.upper(), meta["codigo"], meta["clasificacion"], meta["nombre"], condicion, subtipo))
    row = [year, dz.upper(), meta["codigo"], meta["clasificacion"], meta["nombre"], condicion, subtipo, archivo, ruta]
    rows_by_category.setdefault(categoria, []).append(row)


def generar_reporte_documental_anual(dz: str, year: int | str, maestra_xlsx: Path = STATIONS_XLSX_DEFAULT) -> Path:
    """Genera un Excel anual con documentos formales/administrativos por estación.

    No incluye rutas ni mantenimientos operativos.
    """
    from openpyxl import Workbook

    year = int(year)
    dz = dz.upper()
    base_year = year_dir(dz, year)
    if not base_year.exists():
        raise FileNotFoundError(f"No existe la carpeta del año: {base_year}")

    try:
        catalog = load_station_catalog(maestra_xlsx)
    except Exception:
        catalog = {}
    catalog_by_folder = {v.get("folder_name", ""): v for v in catalog.values()}

    rows_by_category: dict[str, list[list]] = {cat: [] for cat in DOCUMENTAL_CATEGORIES}
    resumen_keys: set = set()

    for station_path in sorted([p for p in base_year.iterdir() if _is_station_folder(p)]):
        meta = _station_meta_from_folder(station_path.name, catalog_by_folder)

        # Carpetas documentales por estación.
        for categoria in [c for c in DOCUMENTAL_CATEGORIES if c != "CONVENIOS"]:
            cdir = station_path / categoria
            if not cdir.exists():
                continue

            if categoria == "SINIESTROS":
                for f in sorted([p for p in cdir.rglob("*") if p.is_file() and not p.name.startswith("legajo_index")]):
                    try:
                        rel_to_cat = f.relative_to(cdir)
                        first = rel_to_cat.parts[0] if len(rel_to_cat.parts) > 1 else ""
                    except Exception:
                        first = ""
                    subtipo = first if first in SINIESTRO_SUBTIPOS else "OTRO"
                    ruta = os.path.relpath(f, start=base_year)
                    _append_doc_rows(rows_by_category, resumen_keys, year, dz, meta, categoria, f.name, ruta, subtipo)
            else:
                for f in sorted([p for p in cdir.rglob("*") if p.is_file() and not p.name.startswith("legajo_index")]):
                    ruta = os.path.relpath(f, start=base_year)
                    _append_doc_rows(rows_by_category, resumen_keys, year, dz, meta, categoria, f.name, ruta, "")

        # Convenios referenciados en el índice anual de la estación.
        index_path = station_path / "legajo_index.xlsx"
        for row in read_xlsx_as_dicts(index_path):
            cat = str(row.get("categoria", "") or "").upper().strip()
            if cat != "CONVENIOS":
                continue
            row_year = str(row.get("year", "") or "").strip()
            if row_year and row_year != f"{year:04d}" and row_year != str(year):
                continue
            archivo = str(row.get("filename", "") or "").strip()
            ruta = str(row.get("relpath", "") or "").strip()
            _append_doc_rows(rows_by_category, resumen_keys, year, dz, meta, "CONVENIOS", archivo, ruta, "")

    resumen_rows = [list(k) for k in sorted(resumen_keys, key=lambda x: (x[1], x[2], x[5], x[6]))]
    detalle_rows = []
    for cat in DOCUMENTAL_CATEGORIES:
        detalle_rows.extend(rows_by_category.get(cat, []))
    detalle_rows.sort(key=lambda x: (x[1], x[2], x[5], x[6], x[7]))

    outdir = reporte_documental_anual_dir(dz, year)
    ensure_dirs(outdir)
    outpath = reporte_documental_anual_path(dz, year)

    wb = Workbook()
    ws = wb.active
    ws.title = "RESUMEN"
    ws.append(["Año", "DZ", "Código", "Clasificación", "Estación", "Condición documental", "Subtipo documental"])
    for row in resumen_rows:
        ws.append(row)
    _apply_xlsx_style(ws)

    ws = wb.create_sheet("DETALLE")
    ws.append(["Año", "DZ", "Código", "Clasificación", "Estación", "Condición documental", "Subtipo documental", "Archivo", "Ruta"])
    for row in detalle_rows:
        ws.append(row)
    _apply_xlsx_style(ws)

    sheet_names = {
        "INSTALACION_ESTACION": "INST_ESTACION",
        "CONVENIOS": "CONVENIOS",
        "CESE_OBSERVADOR": "CESE_OBSERVADOR",
        "SINIESTROS": "SINIESTROS",
        "REUBICACION": "REUBICACION",
        "SOLICITUD_REUBICACION": "SOLICITUD_REUBICACION",
        "BAJA_ESTACION": "BAJA_ESTACION",
        "ACTUALIZACION_METADATA": "ACT_METADATA",
    }
    # Crea pestañas específicas solo cuando existe al menos un archivo de esa sección.
    # RESUMEN siempre queda visible como hoja principal del reporte.
    for cat in DOCUMENTAL_CATEGORIES:
        cat_rows = rows_by_category.get(cat, [])
        if not cat_rows:
            continue
        ws = wb.create_sheet(sheet_names.get(cat, cat)[:31])
        ws.append(["Año", "DZ", "Código", "Clasificación", "Estación", "Condición documental", "Subtipo documental", "Archivo", "Ruta"])
        for row in sorted(cat_rows, key=lambda x: (x[2], x[6], x[7])):
            ws.append(row)
        _apply_xlsx_style(ws)

    wb.save(outpath)
    return outpath

# =========================
# INIT
# =========================
def init_structure(dz: str, years=(2024, 2025), codigos_estaciones=None, maestra_xlsx: Path = STATIONS_XLSX_DEFAULT, ficha_overwrite: bool = False):
    ensure_dirs(ROOT)
    dz_root = dz_dir(dz)
    ensure_dirs(dz_root)
    ensure_dirs(dz_root / DZ_CONVENIOS_DIRNAME)
    ensure_dz_convenios_indices(dz)
    ensure_dirs(dz_ficha_matricula_dir(dz))

    try:
        ficha_path = create_ficha_excel(dz, overwrite=ficha_overwrite)
        print(f"   📘 Excel de Ficha de Matricula listo: {ficha_path}")
    except RuntimeError as e:
        print(f"   ⚠️ No se pudo crear el Excel de Ficha de Matricula: {e}")

    write_xlsx_if_missing(estaciones_rutas_path(dz), ["DZ","Estacion","Ruta","Vigente_desde","Vigente_hasta"])

    estaciones_meta = [get_station_meta(codigo, maestra_xlsx) for codigo in (codigos_estaciones or [])]
    for y in years:
        ydir = year_dir(dz, y)
        ensure_dirs(ydir)
        for meta in estaciones_meta:
            ed = estacion_dir(dz, y, meta)
            ensure_dirs(ed)
            for cat in sorted(set(CATEGORIAS_ESTACION)):
                ensure_dirs(ed / cat)
            ensure_dirs(mantenimiento_correctivo_dir_estacion(dz, y, meta))
            idx_legacy = ed / "legajo_index.xlsx"
            write_xlsx_if_missing(idx_legacy, ["categoria","year","dz","estacion_slug","filename","relpath"])
            idx_explicit = ed / f"legajo_index_{meta['folder_name']}_{int(y):04d}.xlsx"
            if not idx_explicit.exists():
                shutil.copy2(idx_legacy, idx_explicit)
            write_estacion_readme(dz, y, meta)
        mgdir = mantenimiento_grupal_dir(dz, y)
        ensure_dirs(mgdir)
        ensure_mantenimiento_grupal_indices(dz, y)

        rdir = rutas_dir(dz, y)
        ensure_dirs(rdir)
        for t in RUTAS_TIPOS:
            ensure_dirs(rdir / t)
        ensure_rutas_indices(dz, y)

# =========================
# PROMPTS
# =========================
def prompt_categoria_estacion() -> str:
    print("\nCategorías disponibles (por estación):")
    for i, c in enumerate(CATEGORIAS_ESTACION, 1):
        print(f"  {i}. {c}")
    while True:
        raw = input("Elige categoría (número o texto): ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(CATEGORIAS_ESTACION):
            return CATEGORIAS_ESTACION[int(raw)-1]
        upper = raw.upper()
        if upper in CATEGORIAS_ESTACION:
            return upper
        print("⚠️ Valor no válido. Intenta nuevamente.")


def prompt_dz() -> str:
    while True:
        raw = input("DZ (formato DZ01..DZ13): ").strip().upper()
        if re.match(r"^DZ(0[1-9]|1[0-3])$", raw):
            return raw
        print("⚠️ Formato inválido. Ejemplo: DZ06")


def prompt_codigo_estacion(maestra_xlsx: Path) -> str:
    catalog = load_station_catalog(maestra_xlsx)
    while True:
        raw = input("Código de estación: ").strip()
        if raw in catalog:
            print(f"✅ Estación encontrada: {catalog[raw]['folder_name']}")
            return raw
        print(f"⚠️ Código no encontrado en {maestra_xlsx.name}. Intenta nuevamente.")


def prompt_copy() -> bool:
    while True:
        raw = input("¿Copiar en lugar de mover? [s/N]: ").strip().lower()
        if raw in ("s", "si", "sí", "y", "yes"):
            return True
        if raw in ("", "n", "no"):
            return False
        print("⚠️ Responde 's' para copiar o 'n' para mover.")


def prompt_ruta() -> str:
    while True:
        raw = input("Ruta (ej. RUTA_01): ").strip().upper()
        if re.match(r"^RUTA_\d{2}$", raw):
            return raw
        print("⚠️ Formato inválido. Usa RUTA_01, RUTA_02, ...")


def prompt_tipo_ruta() -> str:
    print("\nTipos de RUTA:")
    for i, t in enumerate(RUTAS_TIPOS, 1):
        print(f"  {i}. {t}")
    while True:
        raw = input("Elige tipo (número o texto): ").strip().upper()
        if raw.isdigit() and 1 <= int(raw) <= len(RUTAS_TIPOS):
            return RUTAS_TIPOS[int(raw)-1]
        if raw in RUTAS_TIPOS:
            return raw
        print("⚠️ Valor no válido. Intenta nuevamente.")


def prompt_codigos_estaciones(maestra_xlsx: Path) -> list[str]:
    catalog = load_station_catalog(maestra_xlsx)
    while True:
        raw = input("Códigos de estaciones incluidos (separados por coma, o vacío si no aplica): ").strip()
        if not raw:
            return []
        codigos = [s.strip() for s in raw.split(",") if s.strip()]
        invalidos = [c for c in codigos if c not in catalog]
        if invalidos:
            print(f"⚠️ Códigos no encontrados en {maestra_xlsx.name}: {', '.join(invalidos)}")
            continue
        return codigos

# =========================
# INTERACTIVE WRAPPERS
# =========================
def interactive_for_estacion(args):
    maestra = Path(args.maestra).expanduser()
    categoria = args.categoria or prompt_categoria_estacion()
    dz = args.dz or prompt_dz()
    codigo = args.codigo or prompt_codigo_estacion(maestra)
    fecha = args.fecha or prompt_fecha_ddmmyyyy()
    copy = args.copy if args.copy is not None else prompt_copy()
    return categoria, dz, codigo, fecha, copy


def interactive_for_ruta(args):
    maestra = Path(args.maestra).expanduser()
    dz = args.dz or prompt_dz()
    ruta = args.ruta or prompt_ruta()
    tipo = args.tipo or prompt_tipo_ruta()
    fecha = args.fecha or prompt_fecha_ddmmyyyy()
    codigos = [s.strip() for s in (args.estaciones or "").split(",") if s.strip()] or prompt_codigos_estaciones(maestra)
    copy = args.copy if args.copy is not None else prompt_copy()
    return dz, ruta, tipo, fecha, codigos, copy


def interactive_for_codigo_estacion(args, requiere_ruta: bool = False):
    maestra = Path(args.maestra).expanduser()
    dz = args.dz or prompt_dz()
    ruta = args.ruta or prompt_ruta() if requiere_ruta else None
    codigo = args.codigo or prompt_codigo_estacion(maestra)
    fecha = args.fecha or prompt_fecha_ddmmyyyy()
    copy = args.copy if args.copy is not None else prompt_copy()
    if requiere_ruta:
        return dz, ruta, codigo, fecha, copy
    return dz, codigo, fecha, copy


def interactive_for_convenio_dz(args):
    maestra = Path(args.maestra).expanduser()
    dz = args.dz or prompt_dz()
    fecha = args.fecha or prompt_fecha_ddmmyyyy()
    codigos = [s.strip() for s in (args.estaciones or "").split(",") if s.strip()] or prompt_codigos_estaciones(maestra)
    copy = args.copy if args.copy is not None else prompt_copy()
    return dz, fecha, codigos, copy

# =========================
# CLI
# =========================
def main():
    ap = argparse.ArgumentParser(description="Gestor de legajos con código+clasificación+nombre. Fechas: DD-MM-YYYY.")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p_init = sub.add_parser("init", help="Crear estructura base para una DZ")
    p_init.add_argument("--dz", required=True, help="Ej: DZ06")
    p_init.add_argument("--years", nargs="+", type=int, default=[2025, 2024])
    p_init.add_argument("--estaciones", default="", help="Códigos de estaciones separados por coma")
    p_init.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")
    p_init.add_argument("--ficha-overwrite", action="store_true", help="Si existe el Excel de Ficha, reemplazarlo")

    p_add = sub.add_parser("add", help="Registrar un informe por ESTACIÓN")
    p_add.add_argument("--src", required=True, help="Archivo fuente")
    p_add.add_argument("--categoria", choices=CATEGORIAS_ESTACION)
    p_add.add_argument("--dz", help="Ej: DZ06")
    p_add.add_argument("--codigo", help="Código de la estación")
    p_add.add_argument("--fecha", help="DD-MM-YYYY")
    p_add.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")
    p_add.add_argument("--copy", action="store_const", const=True, default=None, help="Copiar en lugar de mover")
    p_add.add_argument("--subtipo-siniestro", choices=SINIESTRO_SUBTIPOS, default="OTRO", help="Solo aplica cuando la categoría es SINIESTROS")

    p_mat = sub.add_parser("addmatricula", help="Registrar documentos de INSTALACIÓN por estación")
    p_mat.add_argument("--src", help="Archivo fuente único, mantenido por compatibilidad")
    p_mat.add_argument("--srcs", nargs="+", help="Archivos fuente de matrícula")
    p_mat.add_argument("--tipos", nargs="+", help="Tipos de documento: " + ", ".join(MATRICULA_DOC_TYPES))
    p_mat.add_argument("--nombres-finales", nargs="+", default=None, help="Nombres finales para guardar cada documento")
    p_mat.add_argument("--dz", help="Ej: DZ06")
    p_mat.add_argument("--codigo", help="Código de la estación")
    p_mat.add_argument("--fecha", help="DD-MM-YYYY")
    p_mat.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")
    p_mat.add_argument("--copy", action="store_const", const=True, default=None, help="Copiar en lugar de mover")

    p_addmg = sub.add_parser("addmantenimiento_grupal", help="Registrar un informe de mantenimiento grupal sin ruta")
    p_addmg.add_argument("--src", required=True, help="Archivo fuente")
    p_addmg.add_argument("--dz", help="Ej: DZ06")
    p_addmg.add_argument("--fecha", help="DD-MM-YYYY")
    p_addmg.add_argument("--estaciones", help="Códigos de estaciones incluidos separados por coma")
    p_addmg.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")
    p_addmg.add_argument("--responsable", default="")
    p_addmg.add_argument("--obs", default="")
    p_addmg.add_argument("--copy", action="store_const", const=True, default=None)

    p_addr = sub.add_parser("addruta", help="Registrar un informe por RUTA")
    p_addr.add_argument("--src", required=True, help="Archivo fuente")
    p_addr.add_argument("--dz", help="Ej: DZ06")
    p_addr.add_argument("--ruta", help="Ej: RUTA_01")
    p_addr.add_argument("--tipo", choices=RUTAS_TIPOS)
    p_addr.add_argument("--fecha", help="DD-MM-YYYY")
    p_addr.add_argument("--estaciones", help="Códigos de estaciones incluidos separados por coma (opcional)")
    p_addr.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")
    p_addr.add_argument("--responsable", default="")
    p_addr.add_argument("--obs", default="")
    p_addr.add_argument("--copy", action="store_const", const=True, default=None)

    p_add_dz = sub.add_parser("addconvenio_dz", help="Registrar un convenio general de la DZ")
    p_add_dz.add_argument("--src", required=True)
    p_add_dz.add_argument("--dz")
    p_add_dz.add_argument("--fecha", help="DD-MM-YYYY")
    p_add_dz.add_argument("--estaciones", default="", help="Códigos de estaciones incluidos separados por coma (opcional)")
    p_add_dz.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")
    p_add_dz.add_argument("--obs", default="")
    p_add_dz.add_argument("--copy", action="store_const", const=True, default=None)

    p_addcl = sub.add_parser("addchecklist", help="Agregar un Checklist a MANTENIMIENTO/RUTA_XX")
    p_addcl.add_argument("--src", required=True)
    p_addcl.add_argument("--dz")
    p_addcl.add_argument("--ruta", help="Ej: RUTA_01")
    p_addcl.add_argument("--codigo", help="Código de la estación")
    p_addcl.add_argument("--fecha", help="DD-MM-YYYY")
    p_addcl.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")
    p_addcl.add_argument("--copy", action="store_const", const=True, default=None)

    p_addes = sub.add_parser("addestado_situacional", help="Agregar un Estado situacional en MANTENIMIENTO/RUTA_XX")
    p_addes.add_argument("--src", required=True)
    p_addes.add_argument("--dz")
    p_addes.add_argument("--ruta", help="Ej: RUTA_01")
    p_addes.add_argument("--codigo", help="Código de la estación")
    p_addes.add_argument("--fecha", help="DD-MM-YYYY")
    p_addes.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")
    p_addes.add_argument("--copy", action="store_const", const=True, default=None)

    p_addft = sub.add_parser("addfoto", help="Agregar una foto a MANTENIMIENTO/RUTA_XX")
    p_addft.add_argument("--src", required=True)
    p_addft.add_argument("--dz")
    p_addft.add_argument("--ruta", help="Ej: RUTA_01")
    p_addft.add_argument("--codigo", help="Código de la estación")
    p_addft.add_argument("--fecha", help="DD-MM-YYYY")
    p_addft.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")
    p_addft.add_argument("--copy", action="store_const", const=True, default=None)

    p_fichax = sub.add_parser("addficha_dz", help="Crear Excel plantilla en 'DZ/Ficha de Matricula/'")
    p_fichax.add_argument("--dz", help="Ej: DZ06")
    p_fichax.add_argument("--filename", default="")
    p_fichax.add_argument("--overwrite", action="store_true")

    p_repdoc = sub.add_parser("reporte_documental_anual", help="Generar reporte documental anual de una DZ/año")
    p_repdoc.add_argument("--dz", required=True, help="Ej: DZ04")
    p_repdoc.add_argument("--year", required=True, type=int, help="Ej: 2025")
    p_repdoc.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")

    p_index = sub.add_parser("index", help="Reconstruir índice de una estación en un año")
    p_index.add_argument("--dz", required=True)
    p_index.add_argument("--year", required=True, type=int)
    p_index.add_argument("--codigo", required=True, help="Código de la estación")
    p_index.add_argument("--maestra", default=str(STATIONS_XLSX_DEFAULT), help="Excel maestro de estaciones")

    p_ficha = sub.add_parser("mk_ficha_dz", help="Crear/asegurar la carpeta 'Ficha de Matricula' a nivel de DZ")
    p_ficha.add_argument("--dz", required=True, help="Ej: DZ06")

    args = ap.parse_args()

    if args.cmd == "init":
        codigos = [s.strip() for s in args.estaciones.split(",") if s.strip()]
        init_structure(args.dz, years=tuple(args.years), codigos_estaciones=codigos, maestra_xlsx=Path(args.maestra).expanduser(), ficha_overwrite=args.ficha_overwrite)
        print(f"✅ Estructura base creada: {dz_dir(args.dz)}")

    elif args.cmd == "add":
        src = Path(args.src).expanduser()
        if not src.exists():
            print(f"❌ No existe el archivo fuente: {src}", file=sys.stderr); sys.exit(1)
        categoria, dzv, codigo, fecha, copy = interactive_for_estacion(args)
        dst = add_report_estacion(src, categoria, dzv, codigo, fecha, Path(args.maestra).expanduser(), copy, getattr(args, "subtipo_siniestro", "OTRO"))
        print(f"✅ Informe por estación agregado: {dst}")

    elif args.cmd == "addmatricula":
        src_values = args.srcs or ([args.src] if args.src else [])
        if not src_values:
            print("❌ Debes indicar --srcs archivo1 archivo2 ... o --src archivo", file=sys.stderr); sys.exit(1)

        srcs = [Path(x).expanduser() for x in src_values]
        for src in srcs:
            if not src.exists():
                print(f"❌ No existe el archivo fuente: {src}", file=sys.stderr); sys.exit(1)

        tipos = args.tipos or ["OTRO"] * len(srcs)
        if len(tipos) != len(srcs):
            print("❌ La cantidad de --tipos debe coincidir con la cantidad de --srcs", file=sys.stderr); sys.exit(1)

        if args.nombres_finales is not None and len(args.nombres_finales) != len(srcs):
            print("❌ La cantidad de --nombres-finales debe coincidir con la cantidad de --srcs", file=sys.stderr); sys.exit(1)

        maestra = Path(args.maestra).expanduser()
        dzv = args.dz or prompt_dz()
        codigo = args.codigo or prompt_codigo_estacion(maestra)
        fecha = args.fecha or prompt_fecha_ddmmyyyy()
        copy = args.copy if args.copy is not None else prompt_copy()
        destinos = add_documentos_matricula(srcs, tipos, args.nombres_finales, dzv, codigo, fecha, maestra, copy)
        print("✅ Documentos de matrícula agregados:")
        for dst in destinos:
            print(f"   - {dst}")

    elif args.cmd == "addmantenimiento_grupal":
        src = Path(args.src).expanduser()
        if not src.exists():
            print(f"❌ No existe el archivo fuente: {src}", file=sys.stderr); sys.exit(1)
        maestra = Path(args.maestra).expanduser()
        dzv = args.dz or prompt_dz()
        fecha = args.fecha or prompt_fecha_ddmmyyyy()
        codigos = [x.strip() for x in (args.estaciones or "").split(",") if x.strip()] or prompt_codigos_estaciones(maestra)
        copy = args.copy if args.copy is not None else prompt_copy()
        dst = add_mantenimiento_grupal(src, dzv, fecha, codigos, maestra, args.responsable, args.obs, copy)
        print(f"✅ Mantenimiento grupal sin ruta agregado: {dst}")

    elif args.cmd == "addruta":
        src = Path(args.src).expanduser()
        if not src.exists():
            print(f"❌ No existe el archivo fuente: {src}", file=sys.stderr); sys.exit(1)
        dzv, ruta, tipo, fecha, codigos, copy = interactive_for_ruta(args)
        dst = add_report_ruta(src, dzv, ruta, tipo, fecha, codigos, Path(args.maestra).expanduser(), args.responsable, args.obs, copy)
        print(f"✅ Informe por ruta agregado: {dst}")

    elif args.cmd == "addconvenio_dz":
        src = Path(args.src).expanduser()
        if not src.exists():
            print(f"❌ No existe el archivo fuente: {src}", file=sys.stderr); sys.exit(1)
        dzv, fecha, codigos, copy = interactive_for_convenio_dz(args)
        dst = add_convenio_dz(src, dzv, fecha, codigos, Path(args.maestra).expanduser(), args.obs or "", copy)
        print(f"✅ Convenio general de DZ agregado: {dst}")

    elif args.cmd == "addchecklist":
        src = Path(args.src).expanduser()
        if not src.exists():
            print(f"❌ No existe el archivo fuente: {src}", file=sys.stderr); sys.exit(1)
        dzv, ruta, codigo, fecha, copy = interactive_for_codigo_estacion(args, requiere_ruta=True)
        dst = add_checklist_estacion(src, dzv, ruta, codigo, fecha, Path(args.maestra).expanduser(), copy)
        print(f"✅ Checklist agregado en MANTENIMIENTO/{ruta.upper()}: {dst}")

    elif args.cmd == "addestado_situacional":
        src = Path(args.src).expanduser()
        if not src.exists():
            print(f"❌ No existe el archivo fuente: {src}", file=sys.stderr); sys.exit(1)
        dzv, ruta, codigo, fecha, copy = interactive_for_codigo_estacion(args, requiere_ruta=True)
        dst = add_estado_situacional_estacion(src, dzv, ruta, codigo, fecha, Path(args.maestra).expanduser(), copy)
        print(f"✅ ESTADO SITUACIONAL agregado en MANTENIMIENTO/{ruta.upper()}: {dst}")

    elif args.cmd == "addfoto":
        src = Path(args.src).expanduser()
        if not src.exists():
            print(f"❌ No existe el archivo fuente: {src}", file=sys.stderr); sys.exit(1)
        dzv, ruta, codigo, fecha, copy = interactive_for_codigo_estacion(args, requiere_ruta=True)
        dst = add_foto_estacion(src, dzv, ruta, codigo, fecha, Path(args.maestra).expanduser(), copy)
        print(f"✅ Foto agregada en MANTENIMIENTO/{ruta.upper()}: {dst}")

    elif args.cmd == "addficha_dz":
        dzv = args.dz or prompt_dz()
        fname = args.filename.strip() or build_filename_ficha(dzv)
        try:
            out = create_ficha_excel(dzv, filename=fname, overwrite=args.overwrite)
            print(f"✅ Excel de Ficha de Matricula creado: {out}")
            print("   Hoja: MATRÍCULA | Columnas:", ", ".join(FICHA_COLUMNS))
        except RuntimeError as e:
            print(f"❌ {e}", file=sys.stderr); sys.exit(1)

    elif args.cmd == "reporte_documental_anual":
        out = generar_reporte_documental_anual(args.dz, args.year, Path(args.maestra).expanduser())
        print(f"✅ Reporte documental anual generado: {out}")

    elif args.cmd == "index":
        meta = get_station_meta(args.codigo, Path(args.maestra).expanduser())
        write_estacion_readme(args.dz, args.year, meta)
        update_legajo_index_estacion(args.dz, args.year, meta)
        base = estacion_dir(args.dz, args.year, meta)
        explicit_idx = base / f"legajo_index_{meta['folder_name']}_{int(args.year):04d}.xlsx"
        print("✅ Índices reconstruidos:")
        print(f"   - {base / 'legajo_index.xlsx'}")
        print(f"   - {explicit_idx}")

    elif args.cmd == "mk_ficha_dz":
        target = dz_ficha_matricula_dir(args.dz)
        ensure_dirs(target)
        print(f"✅ Carpeta creada/asegurada: {target}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        sys.exit(1)